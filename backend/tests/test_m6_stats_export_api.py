import datetime as dt
import json
from io import BytesIO
from urllib.parse import parse_qs, urlparse

import pytest
from django.contrib.auth.models import User
from django.core.cache import cache
from django.http import QueryDict
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.audit.models import AuditLog
from apps.changesets.models import ChangeEntry, ChangeSet
from apps.schemas.models import DataSchema, SchemaVersion, TableCollaborator
from apps.stats import api as stats_api
from apps.stats import views as stats_views
from apps.temporal.models import Entity, TemporalRecord


@pytest.fixture
def users(db):
    return {
        "owner": User.objects.create_user(username="owner", password="pass"),
        "viewer": User.objects.create_user(username="viewer", password="pass"),
        "outsider": User.objects.create_user(username="outsider", password="pass"),
    }


@pytest.fixture
def client():
    return APIClient()


def auth(client, user):
    client.force_authenticate(user=user)
    return client


def parsed_query(url: str) -> dict[str, list[str]]:
    return parse_qs(urlparse(url).query, keep_blank_values=True)


@pytest.fixture
def schema(users):
    schema = DataSchema.objects.create(
        schema_code="assets",
        name="Assets",
        description="Asset inventory",
        icon="boxes",
        temporal_mode="continuous",
        identity_field_key="asset_no",
        fields_config=[
            {"key": "asset_no", "label": "Asset No", "type": "text", "introduced_in_version": 1},
            {
                "key": "status",
                "label": "Status",
                "type": "enum",
                "validators": {"options": ["In Use", "Repair", "Retired"]},
                "introduced_in_version": 1,
            },
            {"key": "owner", "label": "Owner", "type": "text", "introduced_in_version": 1},
        ],
        current_version=1,
        owner=users["owner"],
        visibility=DataSchema.Visibility.SHARED,
        created_by=users["owner"],
    )
    SchemaVersion.objects.create(
        schema=schema,
        version=1,
        fields_config=schema.fields_config,
        changelog="Initial version",
        created_by=users["owner"],
    )
    TableCollaborator.objects.create(
        schema=schema,
        user=users["viewer"],
        role=TableCollaborator.Role.VIEWER,
        added_by=users["owner"],
    )
    return schema


@pytest.fixture
def records(schema, users):
    change_set = ChangeSet.objects.create(
        schema=schema,
        summary="Initial and June changes",
        status=ChangeSet.Status.APPLIED,
        created_by=users["owner"],
        applied_at=timezone.make_aware(dt.datetime(2024, 6, 21, 10, 0, 0)),
    )
    asset_a = Entity.objects.create(schema=schema, business_code="A-001", created_by=users["owner"])
    asset_b = Entity.objects.create(schema=schema, business_code="B-001", created_by=users["owner"])
    asset_c = Entity.objects.create(schema=schema, business_code="C-001", created_by=users["owner"])
    record_a1 = TemporalRecord.objects.create(
        entity=asset_a,
        schema_version=1,
        data_payload={"asset_no": "A-001", "status": "In Use", "owner": "Alice"},
        valid_from=dt.date(2024, 1, 1),
        valid_to=dt.date(2024, 6, 5),
        change_set=change_set,
        recorded_by=users["owner"],
    )
    record_a2 = TemporalRecord.objects.create(
        entity=asset_a,
        schema_version=1,
        data_payload={"asset_no": "A-001", "status": "Repair", "owner": "Alice"},
        valid_from=dt.date(2024, 6, 5),
        change_set=change_set,
        recorded_by=users["owner"],
    )
    record_b = TemporalRecord.objects.create(
        entity=asset_b,
        schema_version=1,
        data_payload={"asset_no": "B-001", "status": "In Use", "owner": "Bob"},
        valid_from=dt.date(2024, 3, 1),
        change_set=change_set,
        recorded_by=users["owner"],
    )
    record_c = TemporalRecord.objects.create(
        entity=asset_c,
        schema_version=1,
        data_payload={"asset_no": "C-001", "status": "Retired", "owner": "Cara"},
        valid_from=dt.date(2024, 5, 1),
        valid_to=dt.date(2024, 6, 20),
        change_set=change_set,
        recorded_by=users["owner"],
    )
    for entity, record in ((asset_a, record_a1), (asset_b, record_b), (asset_c, record_c)):
        ChangeEntry.objects.create(
            change_set=change_set,
            entity=entity,
            action=ChangeEntry.Action.CREATE,
            data_after=record.data_payload,
            valid_from=record.valid_from,
            valid_to=record.valid_to,
            new_record=record,
        )
    ChangeEntry.objects.create(
        change_set=change_set,
        entity=asset_a,
        action=ChangeEntry.Action.UPDATE,
        data_before=record_a1.data_payload,
        data_after=record_a2.data_payload,
        valid_from=record_a2.valid_from,
        new_record=record_a2,
    )
    ChangeEntry.objects.create(
        change_set=change_set,
        entity=asset_c,
        action=ChangeEntry.Action.TERMINATE,
        data_before=record_c.data_payload,
        valid_from=dt.date(2024, 6, 20),
    )
    return {"change_set": change_set, "asset_a": asset_a, "asset_b": asset_b, "asset_c": asset_c}


@pytest.mark.django_db
def test_stats_summary_trend_and_distribution(client, users, schema, records):
    summary = auth(client, users["viewer"]).get(
        f"/api/v1/schemas/{schema.id}/stats/summary",
        {"at": "2024-06-30"},
    )
    trend = client.get(
        f"/api/v1/schemas/{schema.id}/stats/trend",
        {"at": "2024-06-30", "range": "3"},
    )
    distribution = client.get(
        f"/api/v1/schemas/{schema.id}/stats/distribution",
        {"at": "2024-06-30", "field": "status"},
    )

    assert summary.status_code == 200
    assert summary.json()["metrics"] == {
        "total": 2,
        "month_created": 0,
        "month_updated": 1,
        "month_terminated": 1,
    }
    assert summary.json()["latest_change_set_id"] == records["change_set"].id

    assert trend.status_code == 200
    assert trend.json()["unit"] == "month"
    assert trend.json()["points"] == [
        {"at": "2024-04-30", "count": 2},
        {"at": "2024-05-31", "count": 3},
        {"at": "2024-06-30", "count": 2},
    ]

    assert distribution.status_code == 200
    assert distribution.json()["field"] == {"key": "status", "label": "Status", "type": "enum"}
    assert distribution.json()["buckets"] == [
        {"value": "In Use", "count": 1},
        {"value": "Repair", "count": 1},
    ]


@pytest.mark.django_db
def test_stats_fast_path_skips_full_current_view_resolution(
    monkeypatch, users, schema, records
):
    def fail_full_resolution(*args, **kwargs):
        raise AssertionError("stats fast path should not resolve full current view")

    monkeypatch.setattr(stats_api, "resolve_current_view", fail_full_resolution)
    params = QueryDict("at=2024-06-30&field=status")

    summary = stats_api.build_summary_payload(schema, params, users["viewer"])
    distribution = stats_api.build_distribution_payload(schema, params, users["viewer"])

    assert summary["metrics"]["total"] == 2
    assert summary["scope"] == {
        "at": "2024-06-30",
        "retro": False,
        "search": "",
        "ordering": "business_code",
        "change_set": None,
    }
    assert distribution["buckets"] == [
        {"value": "In Use", "count": 1},
        {"value": "Repair", "count": 1},
    ]


@pytest.mark.django_db
def test_stats_use_full_current_filter_instead_of_page_slice(client, users, schema, records):
    change_set = ChangeSet.objects.create(
        schema=schema,
        summary="Add one more active asset",
        status=ChangeSet.Status.APPLIED,
        created_by=users["owner"],
        applied_at=timezone.make_aware(dt.datetime(2024, 6, 22, 10, 0, 0)),
    )
    asset_d = Entity.objects.create(
        schema=schema,
        business_code="D-001",
        created_by=users["owner"],
    )
    record_d = TemporalRecord.objects.create(
        entity=asset_d,
        schema_version=1,
        data_payload={"asset_no": "D-001", "status": "In Use", "owner": "Dora"},
        valid_from=dt.date(2024, 6, 22),
        change_set=change_set,
        recorded_by=users["owner"],
    )
    ChangeEntry.objects.create(
        change_set=change_set,
        entity=asset_d,
        action=ChangeEntry.Action.CREATE,
        data_after=record_d.data_payload,
        valid_from=record_d.valid_from,
        new_record=record_d,
    )
    params = {
        "at": "2024-06-30",
        "search": "In Use",
        "page": "1",
        "page_size": "1",
        "field": "status",
    }
    api = auth(client, users["viewer"])

    summary = api.get(f"/api/v1/schemas/{schema.id}/stats/summary", params)
    distribution = api.get(f"/api/v1/schemas/{schema.id}/stats/distribution", params)

    assert summary.status_code == 200, summary.json()
    assert summary.json()["scope"] == {
        "at": "2024-06-30",
        "retro": False,
        "search": "In Use",
        "ordering": "business_code",
        "change_set": None,
    }
    assert summary.json()["metrics"]["total"] == 2
    assert distribution.status_code == 200, distribution.json()
    assert distribution.json()["scope"] == summary.json()["scope"]
    assert distribution.json()["buckets"] == [{"value": "In Use", "count": 2}]


@pytest.mark.django_db
def test_stats_trend_auto_uses_day_unit_for_short_history(client, users, schema):
    change_set = ChangeSet.objects.create(
        schema=schema,
        summary="Short history",
        status=ChangeSet.Status.APPLIED,
        created_by=users["owner"],
        applied_at=timezone.make_aware(dt.datetime(2024, 6, 18, 10, 0, 0)),
    )
    entity = Entity.objects.create(schema=schema, business_code="A-001", created_by=users["owner"])
    record = TemporalRecord.objects.create(
        entity=entity,
        schema_version=1,
        data_payload={"asset_no": "A-001", "status": "In Use", "owner": "Alice"},
        valid_from=dt.date(2024, 6, 18),
        change_set=change_set,
        recorded_by=users["owner"],
    )
    ChangeEntry.objects.create(
        change_set=change_set,
        entity=entity,
        action=ChangeEntry.Action.CREATE,
        data_after=record.data_payload,
        valid_from=record.valid_from,
        new_record=record,
    )

    response = auth(client, users["viewer"]).get(
        f"/api/v1/schemas/{schema.id}/stats/trend",
        {"at": "2024-06-20", "unit": "auto"},
    )

    assert response.status_code == 200, response.json()
    assert response.json()["unit"] == "day"
    assert response.json()["range"] == 7
    assert response.json()["points"] == [
        {"at": "2024-06-14", "count": 0},
        {"at": "2024-06-15", "count": 0},
        {"at": "2024-06-16", "count": 0},
        {"at": "2024-06-17", "count": 0},
        {"at": "2024-06-18", "count": 1},
        {"at": "2024-06-19", "count": 1},
        {"at": "2024-06-20", "count": 1},
    ]


@pytest.mark.django_db
def test_stats_distribution_flattens_multi_enum_values(client, users):
    schema = DataSchema.objects.create(
        schema_code="tagged_assets",
        name="Tagged Assets",
        description="Assets with tags",
        icon="boxes",
        temporal_mode="continuous",
        identity_field_key="asset_no",
        fields_config=[
            {"key": "asset_no", "label": "Asset No", "type": "text", "introduced_in_version": 1},
            {
                "key": "tags",
                "label": "Tags",
                "type": "multi-enum",
                "validators": {"options": ["Core", "Lab", "Retired"]},
                "introduced_in_version": 1,
            },
        ],
        current_version=1,
        owner=users["owner"],
        visibility=DataSchema.Visibility.PRIVATE,
        created_by=users["owner"],
    )
    SchemaVersion.objects.create(
        schema=schema,
        version=1,
        fields_config=schema.fields_config,
        changelog="Initial version",
        created_by=users["owner"],
    )
    change_set = ChangeSet.objects.create(
        schema=schema,
        summary="Initial tagged assets",
        status=ChangeSet.Status.APPLIED,
        created_by=users["owner"],
        applied_at=timezone.make_aware(dt.datetime(2024, 6, 1, 10, 0, 0)),
    )
    asset_a = Entity.objects.create(schema=schema, business_code="A-001", created_by=users["owner"])
    asset_b = Entity.objects.create(schema=schema, business_code="B-001", created_by=users["owner"])
    TemporalRecord.objects.create(
        entity=asset_a,
        schema_version=1,
        data_payload={"asset_no": "A-001", "tags": ["Core", "Lab"]},
        valid_from=dt.date(2024, 6, 1),
        change_set=change_set,
        recorded_by=users["owner"],
    )
    TemporalRecord.objects.create(
        entity=asset_b,
        schema_version=1,
        data_payload={"asset_no": "B-001", "tags": ["Lab"]},
        valid_from=dt.date(2024, 6, 1),
        change_set=change_set,
        recorded_by=users["owner"],
    )

    response = auth(client, users["owner"]).get(
        f"/api/v1/schemas/{schema.id}/stats/distribution",
        {"at": "2024-06-30", "field": "tags"},
    )

    assert response.status_code == 200, response.json()
    assert response.json()["field"] == {"key": "tags", "label": "Tags", "type": "multi-enum"}
    assert response.json()["buckets"] == [
        {"value": "Core", "count": 1},
        {"value": "Lab", "count": 2},
    ]


@pytest.mark.django_db
def test_stats_distribution_rejects_invalid_field_params(client, users, schema, records):
    schema.fields_config = [
        *schema.fields_config,
        {
            "key": "legacy_status",
            "label": "Legacy Status",
            "type": "enum",
            "validators": {"options": ["Old"]},
            "introduced_in_version": 1,
            "deprecated": True,
        },
    ]
    schema.save(update_fields=["fields_config"])
    api = auth(client, users["viewer"])

    non_distributable = api.get(
        f"/api/v1/schemas/{schema.id}/stats/distribution",
        {"at": "2024-06-30", "field": "owner"},
    )
    missing = api.get(
        f"/api/v1/schemas/{schema.id}/stats/distribution",
        {"at": "2024-06-30", "field": "missing"},
    )
    deprecated = api.get(
        f"/api/v1/schemas/{schema.id}/stats/distribution",
        {"at": "2024-06-30", "field": "legacy_status"},
    )

    assert non_distributable.status_code == 400
    assert non_distributable.json() == {"field": "field is not distributable"}
    assert missing.status_code == 400
    assert missing.json() == {"field": "field does not exist"}
    assert deprecated.status_code == 400
    assert deprecated.json() == {"field": "field does not exist"}

    schema.fields_config = [
        field
        for field in schema.fields_config
        if field["type"] not in {"enum", "multi-enum", "boolean"}
    ]
    schema.save(update_fields=["fields_config"])
    no_default = api.get(
        f"/api/v1/schemas/{schema.id}/stats/distribution",
        {"at": "2024-06-30"},
    )

    assert no_default.status_code == 400
    assert no_default.json() == {"field": "no distributable field"}


@pytest.mark.django_db
def test_stats_distribution_hides_masked_fields_from_explicit_and_default_selection(
    client, users, schema, records
):
    schema.fields_config = [
        {
            "key": "asset_no",
            "label": "Asset No",
            "type": "text",
            "introduced_in_version": 1,
        },
        {
            "key": "status",
            "label": "Status",
            "type": "enum",
            "validators": {"options": ["In Use", "Repair", "Retired"]},
            "introduced_in_version": 1,
            "sensitive": True,
            "masking": {"visible_roles": ["owner"]},
        },
        {
            "key": "visibility",
            "label": "Visibility",
            "type": "enum",
            "validators": {"options": ["Public", "Internal"]},
            "introduced_in_version": 1,
        },
        {"key": "owner", "label": "Owner", "type": "text", "introduced_in_version": 1},
    ]
    schema.save(update_fields=["fields_config"])
    for record in TemporalRecord.objects.filter(entity__schema=schema):
        payload = dict(record.data_payload)
        payload["visibility"] = "Internal" if payload.get("owner") == "Bob" else "Public"
        record.data_payload = payload
        record.save(update_fields=["data_payload"])

    viewer_api = auth(client, users["viewer"])
    explicit_fast = viewer_api.get(
        f"/api/v1/schemas/{schema.id}/stats/distribution",
        {"at": "2024-06-30", "field": "status"},
    )
    explicit_fallback = viewer_api.get(
        f"/api/v1/schemas/{schema.id}/stats/distribution",
        {"at": "2024-06-30", "field": "status", "search": "Repair"},
    )
    default_distribution = viewer_api.get(
        f"/api/v1/schemas/{schema.id}/stats/distribution",
        {"at": "2024-06-30"},
    )

    assert explicit_fast.status_code == 400
    assert explicit_fast.json() == {"field": "field does not exist"}
    assert explicit_fallback.status_code == 400
    assert explicit_fallback.json() == {"field": "field does not exist"}
    assert default_distribution.status_code == 200, default_distribution.json()
    assert default_distribution.json()["field"] == {
        "key": "visibility",
        "label": "Visibility",
        "type": "enum",
    }
    assert default_distribution.json()["buckets"] == [
        {"value": "Public", "count": 1},
        {"value": "Internal", "count": 1},
    ]

    owner_response = auth(client, users["owner"]).get(
        f"/api/v1/schemas/{schema.id}/stats/distribution",
        {"at": "2024-06-30", "field": "status"},
    )
    assert owner_response.status_code == 200, owner_response.json()
    assert owner_response.json()["buckets"] == [
        {"value": "In Use", "count": 1},
        {"value": "Repair", "count": 1},
    ]


@pytest.mark.django_db
def test_stats_flow_status_dimension_returns_left_to_right_flows(client, users, schema, records):
    response = auth(client, users["viewer"]).get(
        f"/api/v1/schemas/{schema.id}/stats/flow",
        {
            "left_at": "2024-06-01",
            "right_at": "2024-06-30",
            "dimension": "status",
        },
    )

    assert response.status_code == 200, response.json()
    payload = response.json()
    assert payload["dimension"] == {
        "kind": "status",
        "key": "status",
        "label": "Status",
        "type": "enum",
        "multi_value": False,
        "count_mode": "entities",
    }
    assert payload["scope"] == {
        "left_at": "2024-06-01",
        "right_at": "2024-06-30",
        "retro": False,
        "search": "",
        "ordering": "business_code",
    }
    assert payload["summary"] == {
        "left_count": 3,
        "right_count": 2,
        "entity_count": 3,
        "changed_entity_count": 2,
        "entered_count": 0,
        "exited_count": 1,
        "unchanged_count": 1,
        "flow_count": 3,
        "top_flow": {"from": "In Use", "to": "Repair", "value": 1},
    }
    assert payload["nodes"] == [
        {"id": "left:In Use", "name": "In Use", "side": "left", "value": "In Use", "count": 2},
        {"id": "left:Retired", "name": "Retired", "side": "left", "value": "Retired", "count": 1},
        {"id": "right:Repair", "name": "Repair", "side": "right", "value": "Repair", "count": 1},
        {"id": "right:In Use", "name": "In Use", "side": "right", "value": "In Use", "count": 1},
        {"id": "right:(\u65e0\u503c)", "name": "(\u65e0\u503c)", "side": "right", "value": "(\u65e0\u503c)", "count": 1},
    ]
    assert payload["links"] == [
        {
            "source": "left:In Use",
            "target": "right:Repair",
            "value": 1,
            "from": "In Use",
            "to": "Repair",
            "changed": True,
            "sample_entity_ids": [records["asset_a"].id],
            "snapshot_diff_to": (
                f"/schemas/{schema.id}/diff-studio?"
                "mode=snapshot&left_at=2024-06-01&right_at=2024-06-30&retro=false"
                "&search=&ordering=business_code&flow_dimension=status&flow_from=In+Use"
                "&flow_to=Repair&page=1"
            ),
        },
        {
            "source": "left:In Use",
            "target": "right:In Use",
            "value": 1,
            "from": "In Use",
            "to": "In Use",
            "changed": False,
            "sample_entity_ids": [records["asset_b"].id],
            "snapshot_diff_to": (
                f"/schemas/{schema.id}/diff-studio?"
                "mode=snapshot&left_at=2024-06-01&right_at=2024-06-30&retro=false"
                "&search=&ordering=business_code&flow_dimension=status&flow_from=In+Use"
                "&flow_to=In+Use&page=1"
            ),
        },
        {
            "source": "left:Retired",
            "target": "right:(\u65e0\u503c)",
            "value": 1,
            "from": "Retired",
            "to": "(\u65e0\u503c)",
            "changed": True,
            "sample_entity_ids": [records["asset_c"].id],
            "snapshot_diff_to": (
                f"/schemas/{schema.id}/diff-studio?"
                "mode=snapshot&left_at=2024-06-01&right_at=2024-06-30&retro=false"
                "&search=&ordering=business_code&flow_dimension=status&flow_from=Retired"
                "&flow_to=&page=1"
            ),
        },
    ]
    assert parsed_query(payload["snapshot_diff_to"]) == {
        "mode": ["snapshot"],
        "left_at": ["2024-06-01"],
        "right_at": ["2024-06-30"],
        "retro": ["false"],
        "search": [""],
        "ordering": ["business_code"],
        "flow_dimension": ["status"],
        "page": ["1"],
    }
    first_link_query = parsed_query(payload["links"][0]["snapshot_diff_to"])
    assert first_link_query["flow_dimension"] == ["status"]
    assert first_link_query["flow_from"] == ["In Use"]
    assert first_link_query["flow_to"] == ["Repair"]
    assert payload["heat"] == [
        {"at": "2024-06-05", "count": 1},
        {"at": "2024-06-20", "count": 1},
    ]


@pytest.mark.django_db
def test_stats_flow_labels_dimension_uses_field_level_multi_enum_set_semantics(client, users):
    schema = DataSchema.objects.create(
        schema_code="private_tagged_assets",
        name="Private Tagged Assets",
        description="Assets with tags for flow stats",
        icon="boxes",
        temporal_mode="continuous",
        identity_field_key="asset_no",
        fields_config=[
            {"key": "asset_no", "label": "Asset No", "type": "text", "introduced_in_version": 1},
            {
                "key": "tags",
                "label": "\u6807\u7b7e",
                "type": "multi-enum",
                "validators": {"options": ["Core", "Lab", "Retired"]},
                "introduced_in_version": 1,
            },
        ],
        current_version=1,
        owner=users["owner"],
        visibility=DataSchema.Visibility.PRIVATE,
        created_by=users["owner"],
    )
    SchemaVersion.objects.create(
        schema=schema,
        version=1,
        fields_config=schema.fields_config,
        changelog="Initial version",
        created_by=users["owner"],
    )
    create_set = ChangeSet.objects.create(
        schema=schema,
        summary="Create tagged asset",
        status=ChangeSet.Status.APPLIED,
        created_by=users["owner"],
        applied_at=timezone.make_aware(dt.datetime(2024, 6, 1, 9, 0, 0)),
    )
    update_set = ChangeSet.objects.create(
        schema=schema,
        summary="Update tags",
        status=ChangeSet.Status.APPLIED,
        created_by=users["owner"],
        applied_at=timezone.make_aware(dt.datetime(2024, 6, 18, 9, 0, 0)),
    )
    entity = Entity.objects.create(schema=schema, business_code="A-001", created_by=users["owner"])
    left_record = TemporalRecord.objects.create(
        entity=entity,
        schema_version=1,
        data_payload={"asset_no": "A-001", "tags": ["Core", "Lab"]},
        valid_from=dt.date(2024, 6, 1),
        valid_to=dt.date(2024, 6, 18),
        change_set=create_set,
        recorded_by=users["owner"],
    )
    right_record = TemporalRecord.objects.create(
        entity=entity,
        schema_version=1,
        data_payload={"asset_no": "A-001", "tags": ["Lab", "Retired"]},
        valid_from=dt.date(2024, 6, 18),
        change_set=update_set,
        recorded_by=users["owner"],
    )
    ChangeEntry.objects.create(
        change_set=create_set,
        entity=entity,
        action=ChangeEntry.Action.CREATE,
        data_after=left_record.data_payload,
        valid_from=left_record.valid_from,
        valid_to=left_record.valid_to,
        new_record=left_record,
    )
    ChangeEntry.objects.create(
        change_set=update_set,
        entity=entity,
        action=ChangeEntry.Action.UPDATE,
        data_before=left_record.data_payload,
        data_after=right_record.data_payload,
        valid_from=right_record.valid_from,
        new_record=right_record,
    )

    response = auth(client, users["owner"]).get(
        f"/api/v1/schemas/{schema.id}/stats/flow",
        {
            "left_at": "2024-06-01",
            "right_at": "2024-06-30",
            "dimension": "labels",
        },
    )

    assert response.status_code == 200, response.json()
    payload = response.json()
    assert payload["dimension"] == {
        "kind": "labels",
        "key": "tags",
        "label": "\u6807\u7b7e",
        "type": "multi-enum",
        "multi_value": True,
        "count_mode": "label_assignments",
    }
    assert [(item["from"], item["to"], item["value"]) for item in payload["links"]] == [
        ("Core", "(\u65e0\u6807\u7b7e)", 1),
        ("Lab", "Lab", 1),
        ("(\u65e0\u6807\u7b7e)", "Retired", 1),
    ]
    assert payload["summary"]["entity_count"] == 1
    assert payload["summary"]["changed_entity_count"] == 1
    assert payload["summary"] == {
        "left_count": 2,
        "right_count": 2,
        "entity_count": 1,
        "changed_entity_count": 1,
        "entered_count": 1,
        "exited_count": 1,
        "unchanged_count": 1,
        "flow_count": 3,
        "top_flow": {"from": "Core", "to": "(\u65e0\u6807\u7b7e)", "value": 1},
    }
    assert payload["nodes"] == [
        {"id": "left:Core", "name": "Core", "side": "left", "value": "Core", "count": 1},
        {"id": "left:Lab", "name": "Lab", "side": "left", "value": "Lab", "count": 1},
        {"id": "left:(\u65e0\u6807\u7b7e)", "name": "(\u65e0\u6807\u7b7e)", "side": "left", "value": "(\u65e0\u6807\u7b7e)", "count": 1},
        {"id": "right:(\u65e0\u6807\u7b7e)", "name": "(\u65e0\u6807\u7b7e)", "side": "right", "value": "(\u65e0\u6807\u7b7e)", "count": 1},
        {"id": "right:Lab", "name": "Lab", "side": "right", "value": "Lab", "count": 1},
        {"id": "right:Retired", "name": "Retired", "side": "right", "value": "Retired", "count": 1},
    ]
    assert payload["links"] == [
        {
            "source": "left:Core",
            "target": "right:(\u65e0\u6807\u7b7e)",
            "value": 1,
            "from": "Core",
            "to": "(\u65e0\u6807\u7b7e)",
            "changed": True,
            "sample_entity_ids": [entity.id],
            "snapshot_diff_to": (
                f"/schemas/{schema.id}/diff-studio?"
                "mode=snapshot&left_at=2024-06-01&right_at=2024-06-30&retro=false"
                "&search=&ordering=business_code&flow_dimension=labels&flow_from=Core"
                "&flow_to=&page=1"
            ),
        },
        {
            "source": "left:Lab",
            "target": "right:Lab",
            "value": 1,
            "from": "Lab",
            "to": "Lab",
            "changed": False,
            "sample_entity_ids": [entity.id],
            "snapshot_diff_to": (
                f"/schemas/{schema.id}/diff-studio?"
                "mode=snapshot&left_at=2024-06-01&right_at=2024-06-30&retro=false"
                "&search=&ordering=business_code&flow_dimension=labels&flow_from=Lab"
                "&flow_to=Lab&page=1"
            ),
        },
        {
            "source": "left:(\u65e0\u6807\u7b7e)",
            "target": "right:Retired",
            "value": 1,
            "from": "(\u65e0\u6807\u7b7e)",
            "to": "Retired",
            "changed": True,
            "sample_entity_ids": [entity.id],
            "snapshot_diff_to": (
                f"/schemas/{schema.id}/diff-studio?"
                "mode=snapshot&left_at=2024-06-01&right_at=2024-06-30&retro=false"
                "&search=&ordering=business_code&flow_dimension=labels&flow_from="
                "&flow_to=Retired&page=1"
            ),
        },
    ]
    assert parsed_query(payload["snapshot_diff_to"]) == {
        "mode": ["snapshot"],
        "left_at": ["2024-06-01"],
        "right_at": ["2024-06-30"],
        "retro": ["false"],
        "search": [""],
        "ordering": ["business_code"],
        "flow_dimension": ["labels"],
        "page": ["1"],
    }


@pytest.mark.django_db
def test_stats_flow_search_filters_entity_scope_without_false_entries(
    client, users, schema, records
):
    response = auth(client, users["viewer"]).get(
        f"/api/v1/schemas/{schema.id}/stats/flow",
        {
            "left_at": "2024-06-01",
            "right_at": "2024-06-30",
            "dimension": "status",
            "search": "Repair",
        },
    )

    assert response.status_code == 200, response.json()
    payload = response.json()
    assert payload["scope"] == {
        "left_at": "2024-06-01",
        "right_at": "2024-06-30",
        "retro": False,
        "search": "Repair",
        "ordering": "business_code",
    }
    assert payload["summary"] == {
        "left_count": 1,
        "right_count": 1,
        "entity_count": 1,
        "changed_entity_count": 1,
        "entered_count": 0,
        "exited_count": 0,
        "unchanged_count": 0,
        "flow_count": 1,
        "top_flow": {"from": "In Use", "to": "Repair", "value": 1},
    }
    assert payload["links"] == [
        {
            "source": "left:In Use",
            "target": "right:Repair",
            "value": 1,
            "from": "In Use",
            "to": "Repair",
            "changed": True,
            "sample_entity_ids": [records["asset_a"].id],
            "snapshot_diff_to": (
                f"/schemas/{schema.id}/diff-studio?"
                "mode=snapshot&left_at=2024-06-01&right_at=2024-06-30&retro=false"
                "&search=Repair&ordering=business_code&flow_dimension=status&flow_from=In+Use"
                "&flow_to=Repair&page=1"
            ),
        }
    ]


@pytest.mark.django_db
def test_stats_flow_dimension_alias_matching_and_label_priority(client, users):
    schema = DataSchema.objects.create(
        schema_code="flow_alias_schema",
        name="Flow Alias Schema",
        description="Schema for flow dimension alias selection",
        icon="boxes",
        temporal_mode="continuous",
        identity_field_key="asset_no",
        fields_config=[
            {"key": "asset_no", "label": "Asset No", "type": "text", "introduced_in_version": 1},
            {
                "key": "state",
                "label": "Lifecycle State",
                "type": "enum",
                "validators": {"options": ["Active", "Repair"]},
                "introduced_in_version": 1,
            },
            {
                "key": "dept",
                "label": "Department",
                "type": "enum",
                "validators": {"options": ["Ops", "QA"]},
                "introduced_in_version": 1,
            },
            {
                "key": "categories",
                "label": "Categories",
                "type": "multi-enum",
                "validators": {"options": ["Cat-A"]},
                "introduced_in_version": 1,
            },
            {
                "key": "tags",
                "label": "Tags",
                "type": "multi-enum",
                "validators": {"options": ["Core", "Lab"]},
                "introduced_in_version": 1,
            },
        ],
        current_version=1,
        owner=users["owner"],
        visibility=DataSchema.Visibility.PRIVATE,
        created_by=users["owner"],
    )
    SchemaVersion.objects.create(
        schema=schema,
        version=1,
        fields_config=schema.fields_config,
        changelog="Initial version",
        created_by=users["owner"],
    )
    api = auth(client, users["owner"])

    status_response = api.get(
        f"/api/v1/schemas/{schema.id}/stats/flow",
        {"left_at": "2024-06-01", "right_at": "2024-06-30", "dimension": "status"},
    )
    department_response = api.get(
        f"/api/v1/schemas/{schema.id}/stats/flow",
        {"left_at": "2024-06-01", "right_at": "2024-06-30", "dimension": "department"},
    )
    labels_response = api.get(
        f"/api/v1/schemas/{schema.id}/stats/flow",
        {"left_at": "2024-06-01", "right_at": "2024-06-30", "dimension": "labels"},
    )

    assert status_response.status_code == 200, status_response.json()
    assert department_response.status_code == 200, department_response.json()
    assert labels_response.status_code == 200, labels_response.json()
    assert status_response.json()["dimension"] == {
        "kind": "status",
        "key": "state",
        "label": "Lifecycle State",
        "type": "enum",
        "multi_value": False,
        "count_mode": "entities",
    }
    assert department_response.json()["dimension"] == {
        "kind": "department",
        "key": "dept",
        "label": "Department",
        "type": "enum",
        "multi_value": False,
        "count_mode": "entities",
    }
    assert labels_response.json()["dimension"] == {
        "kind": "labels",
        "key": "tags",
        "label": "Tags",
        "type": "multi-enum",
        "multi_value": True,
        "count_mode": "label_assignments",
    }


@pytest.mark.django_db
@pytest.mark.parametrize(
    ("dimension", "field_key", "field_label", "field_type", "options"),
    [
        ("status", "status_note", "Status Note", "enum", ["Open", "Closed"]),
        ("department", "org_code", "Organization Code", "enum", ["Ops", "QA"]),
        ("labels", "tagline", "Tagline", "multi-enum", ["A", "B"]),
    ],
)
def test_stats_flow_rejects_partial_alias_matches(
    client,
    users,
    dimension,
    field_key,
    field_label,
    field_type,
    options,
):
    schema = DataSchema.objects.create(
        schema_code=f"partial_alias_{dimension}",
        name="Partial Alias Schema",
        description="Schema for negative alias matching",
        icon="boxes",
        temporal_mode="continuous",
        identity_field_key="asset_no",
        fields_config=[
            {"key": "asset_no", "label": "Asset No", "type": "text", "introduced_in_version": 1},
            {
                "key": field_key,
                "label": field_label,
                "type": field_type,
                "validators": {"options": options},
                "introduced_in_version": 1,
            },
        ],
        current_version=1,
        owner=users["owner"],
        visibility=DataSchema.Visibility.PRIVATE,
        created_by=users["owner"],
    )
    SchemaVersion.objects.create(
        schema=schema,
        version=1,
        fields_config=schema.fields_config,
        changelog="Initial version",
        created_by=users["owner"],
    )

    response = auth(client, users["owner"]).get(
        f"/api/v1/schemas/{schema.id}/stats/flow",
        {"left_at": "2024-06-01", "right_at": "2024-06-30", "dimension": dimension},
    )

    assert response.status_code == 400
    assert response.json() == {"dimension": "dimension field does not exist"}


@pytest.mark.django_db
def test_stats_flow_rejects_invalid_scope_and_missing_dimension(client, users, schema, records):
    api = auth(client, users["viewer"])

    reversed_dates = api.get(
        f"/api/v1/schemas/{schema.id}/stats/flow",
        {
            "left_at": "2024-06-30",
            "right_at": "2024-06-01",
            "dimension": "status",
        },
    )
    missing_dimension = api.get(
        f"/api/v1/schemas/{schema.id}/stats/flow",
        {
            "left_at": "2024-06-01",
            "right_at": "2024-06-30",
        },
    )
    invalid_dimension = api.get(
        f"/api/v1/schemas/{schema.id}/stats/flow",
        {
            "left_at": "2024-06-01",
            "right_at": "2024-06-30",
            "dimension": "owner",
        },
    )

    assert reversed_dates.status_code == 400
    assert reversed_dates.json() == {"right_at": "must be on or after left_at"}
    assert missing_dimension.status_code == 400
    assert missing_dimension.json() == {"dimension": "must be status, department, or labels"}
    assert invalid_dimension.status_code == 400
    assert invalid_dimension.json() == {"dimension": "must be status, department, or labels"}


@pytest.mark.django_db
def test_stats_flow_hides_masked_dimension_fields(client, users, schema, records):
    schema.fields_config = [
        {
            "key": "asset_no",
            "label": "Asset No",
            "type": "text",
            "introduced_in_version": 1,
        },
        {
            "key": "status",
            "label": "Status",
            "type": "enum",
            "validators": {"options": ["In Use", "Repair", "Retired"]},
            "introduced_in_version": 1,
            "sensitive": True,
            "masking": {"visible_roles": ["owner"]},
        },
        {"key": "owner", "label": "Owner", "type": "text", "introduced_in_version": 1},
    ]
    schema.save(update_fields=["fields_config"])

    viewer_response = auth(client, users["viewer"]).get(
        f"/api/v1/schemas/{schema.id}/stats/flow",
        {
            "left_at": "2024-06-01",
            "right_at": "2024-06-30",
            "dimension": "status",
        },
    )
    owner_response = auth(client, users["owner"]).get(
        f"/api/v1/schemas/{schema.id}/stats/flow",
        {
            "left_at": "2024-06-01",
            "right_at": "2024-06-30",
            "dimension": "status",
        },
    )

    assert viewer_response.status_code == 400
    assert viewer_response.json() == {"dimension": "dimension field does not exist"}
    assert owner_response.status_code == 200, owner_response.json()


@pytest.mark.django_db
def test_stats_respect_schema_visibility(client, users, schema, records):
    response = auth(client, users["outsider"]).get(
        f"/api/v1/schemas/{schema.id}/stats/summary",
        {"at": "2024-06-30"},
    )

    assert response.status_code == 404


@pytest.mark.django_db
def test_current_view_export_csv_and_xlsx_include_metadata_and_audit_log(
    client, users, schema, records
):
    csv_response = auth(client, users["viewer"]).get(
        f"/api/v1/schemas/{schema.id}/export/current",
        {"format": "csv", "at": "2024-06-30"},
    )
    xlsx_response = client.get(
        f"/api/v1/schemas/{schema.id}/export/current",
        {
            "format": "xlsx",
            "at": "2024-06-30",
            "search": "A-001",
            "ordering": "-business_code",
            "change_set": records["change_set"].id,
        },
    )

    assert csv_response.status_code == 200, csv_response.json()
    assert csv_response["Content-Type"].startswith("text/csv")
    csv_text = csv_response.content.decode("utf-8-sig")
    assert "display_code,valid_from,valid_to,schema_version,Asset No,Status,Owner" in csv_text
    assert "B-001,2024-03-01,,1,B-001,In Use,Bob" in csv_text

    assert xlsx_response.status_code == 200
    workbook = load_workbook(BytesIO(xlsx_response.content), data_only=True)
    assert workbook.sheetnames == ["data", "metadata"]
    data_sheet = workbook["data"]
    metadata = {
        workbook["metadata"].cell(row=index, column=1).value: workbook["metadata"].cell(
            row=index, column=2
        ).value
        for index in range(1, workbook["metadata"].max_row + 1)
    }
    assert data_sheet.max_row == 2
    assert data_sheet.cell(row=1, column=1).value == "display_code"
    assert metadata["schema_code"] == "assets"
    assert metadata["data_at"] == "2024-06-30"
    assert metadata["row_count"] == 1
    assert metadata["export_scope"] == "current_view"
    assert str(metadata["export_id"]).startswith("EXP-")
    snapshot = json.loads(metadata["query_snapshot"])
    dt.datetime.fromisoformat(snapshot["requested_at"])
    assert snapshot == {
        "schema_id": schema.id,
        "user_id": users["viewer"].id,
        "at": "2024-06-30",
        "retro": False,
        "search": "A-001",
        "ordering": "-business_code",
        "change_set": records["change_set"].id,
        "schema_version": 1,
        "requested_at": snapshot["requested_at"],
        "row_count": 1,
    }

    logs = AuditLog.objects.filter(action="data.export").order_by("id")
    assert logs.count() == 2
    assert logs[0].detail["format"] == "csv"
    assert logs[0].detail["row_count"] == 2
    assert logs[1].detail["format"] == "xlsx"
    assert logs[1].detail["row_count"] == 1
    assert logs[1].detail["query_snapshot"] == snapshot


@pytest.mark.django_db
def test_current_view_export_rejects_invisible_schema(client, users, schema, records):
    response = auth(client, users["outsider"]).get(
        f"/api/v1/schemas/{schema.id}/export/current",
        {"format": "csv", "at": "2024-06-30"},
    )

    assert response.status_code == 404


@pytest.mark.django_db
def test_current_view_export_reauthenticates_before_download(
    client, monkeypatch, users, schema, records
):
    permission_checks = []
    build_calls = []

    def allow_once(user, checked_schema):
        assert user == users["viewer"]
        assert checked_schema == schema
        permission_checks.append(True)
        return len(permission_checks) == 1

    def build_export(*args, **kwargs):
        build_calls.append(True)
        return {
            "content": b"display_code\n",
            "format": "csv",
            "filename": "assets.csv",
            "metadata": {
                "export_id": "EXP-test",
                "format": "csv",
                "row_count": 0,
                "export_scope": "current_view",
                "data_at": "2024-06-30",
            },
        }

    monkeypatch.setattr(stats_views, "can_export_schema", allow_once)
    monkeypatch.setattr(stats_views, "build_current_export", build_export)

    response = auth(client, users["viewer"]).get(
        f"/api/v1/schemas/{schema.id}/export/current",
        {"format": "csv", "at": "2024-06-30"},
    )

    assert response.status_code == 403
    assert permission_checks == [True, True]
    assert build_calls == [True]
    assert not AuditLog.objects.filter(action="data.export").exists()


@pytest.mark.django_db
def test_current_view_export_rejects_duplicate_snapshot_while_running(
    client, monkeypatch, users, schema, records
):
    cache.clear()
    url = f"/api/v1/schemas/{schema.id}/export/current"
    params = {"format": "csv", "at": "2024-06-30", "ordering": "-business_code"}
    build_calls = []
    duplicate_response = {}

    def build_export(*args, **kwargs):
        build_calls.append(True)
        if len(build_calls) == 1:
            duplicate_response["response"] = client.get(url, params)
        return {
            "content": b"display_code\n",
            "format": "csv",
            "filename": "assets.csv",
            "metadata": {
                "export_id": f"EXP-test-{len(build_calls)}",
                "format": "csv",
                "row_count": 0,
                "export_scope": "current_view",
                "data_at": "2024-06-30",
            },
        }

    monkeypatch.setattr(stats_views, "build_current_export", build_export)

    response = auth(client, users["viewer"]).get(url, params)

    assert response.status_code == 200
    assert duplicate_response["response"].status_code == 409
    assert duplicate_response["response"].json() == {
        "detail": "same export snapshot is already running"
    }
    assert build_calls == [True]
    assert AuditLog.objects.filter(action="data.export").count() == 1


@pytest.mark.django_db
def test_changeset_export_xlsx_contains_detail_metadata_and_audit_log(
    client, users, schema, records
):
    response = auth(client, users["viewer"]).get(
        f"/api/v1/changesets/{records['change_set'].id}/export"
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook.sheetnames == ["data", "metadata"]
    data_sheet = workbook["data"]
    metadata = {
        workbook["metadata"].cell(row=index, column=1).value: workbook["metadata"].cell(
            row=index, column=2
        ).value
        for index in range(1, workbook["metadata"].max_row + 1)
    }
    assert data_sheet.cell(row=1, column=1).value == "change_set_id"
    assert data_sheet.cell(row=1, column=4).value == "field"
    assert "status" in [data_sheet.cell(row=row, column=4).value for row in range(2, data_sheet.max_row + 1)]
    assert metadata["export_scope"] == "changeset"
    assert metadata["change_set_id"] == records["change_set"].id
    assert metadata["row_count"] == 5

    log = AuditLog.objects.filter(action="data.export").latest("id")
    assert log.target_type == "changeset"
    assert log.target_id == records["change_set"].id
    assert log.detail["export_scope"] == "changeset"
    assert log.detail["row_count"] == 5


@pytest.mark.django_db
def test_entity_lifecycle_export_xlsx_contains_timeline_metadata_and_audit_log(
    client, users, schema, records
):
    response = auth(client, users["viewer"]).get(f"/api/v1/entities/{records['asset_a'].id}/export")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook.sheetnames == ["data", "metadata"]
    data_sheet = workbook["data"]
    metadata = {
        workbook["metadata"].cell(row=index, column=1).value: workbook["metadata"].cell(
            row=index, column=2
        ).value
        for index in range(1, workbook["metadata"].max_row + 1)
    }
    assert data_sheet.max_row == 3
    assert data_sheet.cell(row=1, column=1).value == "display_code"
    assert data_sheet.cell(row=2, column=1).value == "A-001"
    assert data_sheet.cell(row=3, column=6).value == "Repair"
    assert metadata["export_scope"] == "entity"
    assert metadata["entity_id"] == records["asset_a"].id
    assert metadata["business_code"] == "A-001"
    assert metadata["row_count"] == 2

    log = AuditLog.objects.filter(action="data.export").latest("id")
    assert log.target_type == "entity"
    assert log.target_id == records["asset_a"].id
    assert log.detail["export_scope"] == "entity"
    assert log.detail["row_count"] == 2


@pytest.mark.django_db
def test_dashboard_summary_returns_visible_schema_and_activity_counts(
    client, users, schema, records
):
    archived_schema = DataSchema.objects.create(
        schema_code="archived_assets",
        name="Archived Assets",
        description="Archived asset inventory",
        icon="archive",
        temporal_mode="continuous",
        identity_field_key="asset_no",
        fields_config=[{"key": "asset_no", "label": "Asset No", "type": "text"}],
        current_version=1,
        owner=users["owner"],
        visibility=DataSchema.Visibility.SHARED,
        created_by=users["owner"],
        is_archived=True,
    )
    TableCollaborator.objects.create(
        schema=archived_schema,
        user=users["viewer"],
        role=TableCollaborator.Role.VIEWER,
        added_by=users["owner"],
    )
    ChangeSet.objects.create(
        schema=schema,
        summary="Needs approval",
        status=ChangeSet.Status.SUBMITTED,
        approval_required=True,
        approver=users["viewer"],
        created_by=users["owner"],
    )
    recent = ChangeSet.objects.create(
        schema=schema,
        summary="Recent applied",
        status=ChangeSet.Status.APPLIED,
        created_by=users["owner"],
        applied_at=timezone.now(),
    )
    ChangeEntry.objects.create(
        change_set=recent,
        entity=records["asset_b"],
        action=ChangeEntry.Action.UPDATE,
        data_before={"status": "In Use"},
        data_after={"status": "Repair"},
        valid_from=timezone.localdate(),
    )

    response = auth(client, users["viewer"]).get("/api/v1/dashboard/")

    assert response.status_code == 200
    payload = response.json()
    assert payload["schema_count"] == 1
    assert payload["owned_schema_count"] == 0
    assert payload["shared_schema_count"] == 1
    assert payload["public_schema_count"] == 0
    assert payload["archived_schema_count"] == 1
    assert payload["pending_approval_count"] == 1
    assert payload["recent_change_count"] == 1
    assert payload["active_user_count"] == 1


@pytest.mark.django_db
def test_dashboard_summary_counts_shared_card_by_collaboration(client, users):
    DataSchema.objects.create(
        schema_code="viewer_own_shared",
        name="Viewer Own Shared",
        description="",
        icon="box",
        temporal_mode="continuous",
        identity_field_key="asset_no",
        fields_config=[{"key": "asset_no", "label": "Asset No", "type": "text"}],
        current_version=1,
        owner=users["viewer"],
        visibility=DataSchema.Visibility.SHARED,
        created_by=users["viewer"],
    )
    shared_to_viewer = DataSchema.objects.create(
        schema_code="shared_to_viewer",
        name="Shared To Viewer",
        description="",
        icon="box",
        temporal_mode="continuous",
        identity_field_key="asset_no",
        fields_config=[{"key": "asset_no", "label": "Asset No", "type": "text"}],
        current_version=1,
        owner=users["owner"],
        visibility=DataSchema.Visibility.SHARED,
        created_by=users["owner"],
    )
    DataSchema.objects.create(
        schema_code="viewer_own_public",
        name="Viewer Own Public",
        description="",
        icon="box",
        temporal_mode="continuous",
        identity_field_key="asset_no",
        fields_config=[{"key": "asset_no", "label": "Asset No", "type": "text"}],
        current_version=1,
        owner=users["viewer"],
        visibility=DataSchema.Visibility.PUBLIC,
        created_by=users["viewer"],
    )
    archived_shared = DataSchema.objects.create(
        schema_code="archived_shared_to_viewer",
        name="Archived Shared To Viewer",
        description="",
        icon="archive",
        temporal_mode="continuous",
        identity_field_key="asset_no",
        fields_config=[{"key": "asset_no", "label": "Asset No", "type": "text"}],
        current_version=1,
        owner=users["owner"],
        visibility=DataSchema.Visibility.SHARED,
        created_by=users["owner"],
        is_archived=True,
    )
    for item in (shared_to_viewer, archived_shared):
        TableCollaborator.objects.create(
            schema=item,
            user=users["viewer"],
            role=TableCollaborator.Role.VIEWER,
            added_by=users["owner"],
        )

    response = auth(client, users["viewer"]).get("/api/v1/dashboard/")

    assert response.status_code == 200
    payload = response.json()
    assert payload["schema_count"] == 3
    assert payload["owned_schema_count"] == 2
    assert payload["shared_schema_count"] == 1
    assert payload["public_schema_count"] == 1
    assert payload["archived_schema_count"] == 1


@pytest.mark.django_db
def test_dashboard_summary_counts_superuser_visible_schemas_as_managed(client, users):
    admin = User.objects.create_superuser(
        username="admin",
        email="admin@example.com",
        password="pass",
    )
    DataSchema.objects.create(
        schema_code="owner_shared",
        name="Owner Shared",
        description="",
        icon="box",
        temporal_mode="continuous",
        identity_field_key="asset_no",
        fields_config=[{"key": "asset_no", "label": "Asset No", "type": "text"}],
        current_version=1,
        owner=users["owner"],
        visibility=DataSchema.Visibility.SHARED,
        created_by=users["owner"],
    )
    DataSchema.objects.create(
        schema_code="viewer_public",
        name="Viewer Public",
        description="",
        icon="box",
        temporal_mode="continuous",
        identity_field_key="asset_no",
        fields_config=[{"key": "asset_no", "label": "Asset No", "type": "text"}],
        current_version=1,
        owner=users["viewer"],
        visibility=DataSchema.Visibility.PUBLIC,
        created_by=users["viewer"],
    )
    DataSchema.objects.create(
        schema_code="archived_private",
        name="Archived Private",
        description="",
        icon="archive",
        temporal_mode="continuous",
        identity_field_key="asset_no",
        fields_config=[{"key": "asset_no", "label": "Asset No", "type": "text"}],
        current_version=1,
        owner=users["owner"],
        visibility=DataSchema.Visibility.PRIVATE,
        created_by=users["owner"],
        is_archived=True,
    )

    response = auth(client, admin).get("/api/v1/dashboard/")

    assert response.status_code == 200
    payload = response.json()
    assert payload["schema_count"] == 2
    assert payload["owned_schema_count"] == 2
    assert payload["shared_schema_count"] == 0
    assert payload["public_schema_count"] == 1
    assert payload["archived_schema_count"] == 1


@pytest.mark.django_db
def test_changeset_and_entity_exports_respect_visibility(client, users, schema, records):
    changeset_response = auth(client, users["outsider"]).get(
        f"/api/v1/changesets/{records['change_set'].id}/export"
    )
    entity_response = client.get(f"/api/v1/entities/{records['asset_a'].id}/export")

    assert changeset_response.status_code == 404
    assert entity_response.status_code == 404

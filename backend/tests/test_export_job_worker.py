import datetime as dt
import json
from io import BytesIO, StringIO
from pathlib import Path

import pytest
from django.core.files.base import ContentFile
from django.core.management import call_command
from django.http import QueryDict
from django.utils import timezone
from openpyxl import load_workbook

from apps.changesets.models import ChangeEntry, ChangeSet
from apps.schemas.models import DataSchema, SchemaVersion, TableCollaborator
from apps.stats import export_job_worker
from apps.stats.export_job_worker import process_export_jobs
from apps.stats.export_snapshots import (
    build_current_export_query_snapshot,
    build_current_export_snapshot_key,
)
from apps.stats.management.commands import process_export_jobs as process_export_jobs_command
from apps.stats.models import ExportJob
from apps.temporal.models import Entity, TemporalRecord


@pytest.fixture
def users(db, django_user_model):
    return {
        "owner": django_user_model.objects.create_user(username="owner", password="pass"),
        "viewer": django_user_model.objects.create_user(username="viewer", password="pass"),
    }


@pytest.fixture
def schema(users):
    schema = DataSchema.objects.create(
        schema_code="assets",
        name="Assets",
        description="Asset inventory",
        icon="boxes",
        temporal_mode="continuous",
        identity_field_key="asset_no",
        fields_config=[
            {"key": "asset_no", "label": "Asset No", "type": "text", "introduced_in_version": 1},
            {
                "key": "status",
                "label": "Status",
                "type": "enum",
                "validators": {"options": ["In Use", "Repair", "Retired"]},
                "introduced_in_version": 1,
            },
            {"key": "owner", "label": "Owner", "type": "text", "introduced_in_version": 1},
        ],
        current_version=1,
        owner=users["owner"],
        visibility=DataSchema.Visibility.SHARED,
        created_by=users["owner"],
    )
    SchemaVersion.objects.create(
        schema=schema,
        version=1,
        fields_config=schema.fields_config,
        changelog="Initial version",
        created_by=users["owner"],
    )
    TableCollaborator.objects.create(
        schema=schema,
        user=users["viewer"],
        role=TableCollaborator.Role.VIEWER,
        added_by=users["owner"],
    )
    return schema


@pytest.fixture
def records(schema, users):
    change_set = ChangeSet.objects.create(
        schema=schema,
        summary="Initial and June changes",
        status=ChangeSet.Status.APPLIED,
        created_by=users["owner"],
        applied_at=timezone.make_aware(dt.datetime(2024, 6, 21, 10, 0, 0)),
    )
    asset_a = Entity.objects.create(schema=schema, business_code="A-001", created_by=users["owner"])
    asset_b = Entity.objects.create(schema=schema, business_code="B-001", created_by=users["owner"])
    asset_c = Entity.objects.create(schema=schema, business_code="C-001", created_by=users["owner"])
    record_a1 = TemporalRecord.objects.create(
        entity=asset_a,
        schema_version=1,
        data_payload={"asset_no": "A-001", "status": "In Use", "owner": "Alice"},
        valid_from=dt.date(2024, 1, 1),
        valid_to=dt.date(2024, 6, 5),
        change_set=change_set,
        recorded_by=users["owner"],
    )
    record_a2 = TemporalRecord.objects.create(
        entity=asset_a,
        schema_version=1,
        data_payload={"asset_no": "A-001", "status": "Repair", "owner": "Alice"},
        valid_from=dt.date(2024, 6, 5),
        change_set=change_set,
        recorded_by=users["owner"],
    )
    record_b = TemporalRecord.objects.create(
        entity=asset_b,
        schema_version=1,
        data_payload={"asset_no": "B-001", "status": "In Use", "owner": "Bob"},
        valid_from=dt.date(2024, 3, 1),
        change_set=change_set,
        recorded_by=users["owner"],
    )
    record_c = TemporalRecord.objects.create(
        entity=asset_c,
        schema_version=1,
        data_payload={"asset_no": "C-001", "status": "Retired", "owner": "Cara"},
        valid_from=dt.date(2024, 5, 1),
        valid_to=dt.date(2024, 6, 20),
        change_set=change_set,
        recorded_by=users["owner"],
    )
    for entity, record in ((asset_a, record_a1), (asset_b, record_b), (asset_c, record_c)):
        ChangeEntry.objects.create(
            change_set=change_set,
            entity=entity,
            action=ChangeEntry.Action.CREATE,
            data_after=record.data_payload,
            valid_from=record.valid_from,
            valid_to=record.valid_to,
            new_record=record,
        )
    ChangeEntry.objects.create(
        change_set=change_set,
        entity=asset_a,
        action=ChangeEntry.Action.UPDATE,
        data_before=record_a1.data_payload,
        data_after=record_a2.data_payload,
        valid_from=record_a2.valid_from,
        new_record=record_a2,
    )
    ChangeEntry.objects.create(
        change_set=change_set,
        entity=asset_c,
        action=ChangeEntry.Action.TERMINATE,
        data_before=record_c.data_payload,
        valid_from=dt.date(2024, 6, 20),
    )
    return {"change_set": change_set}


def create_job(schema, user, query_string: str) -> ExportJob:
    query_snapshot = build_current_export_query_snapshot(schema, user, QueryDict(query_string))
    return ExportJob.objects.create(
        owner=user,
        schema=schema,
        export_format=query_snapshot["format"],
        snapshot_key=build_current_export_snapshot_key(query_snapshot),
        query_snapshot=query_snapshot,
        row_count_estimate=2,
    )


def create_job_from_spec(schema, user, spec: dict) -> ExportJob:
    query_snapshot = build_current_export_query_snapshot(schema, user, {"export_spec": spec})
    return ExportJob.objects.create(
        owner=user,
        schema=schema,
        export_format=query_snapshot["format"],
        snapshot_key=build_current_export_snapshot_key(query_snapshot),
        query_snapshot=query_snapshot,
        row_count_estimate=2,
    )


@pytest.mark.django_db
def test_process_export_jobs_completes_queued_csv_job(tmp_path, settings, users, schema, records):
    settings.MEDIA_ROOT = tmp_path
    job = create_job(schema, users["viewer"], "format=csv&at=2024-06-30")

    summary = process_export_jobs(limit=1)

    job.refresh_from_db()
    content = job.file.read()

    assert summary["processed"] == 1
    assert job.status == ExportJob.Status.COMPLETED
    assert content.startswith(b"\xef\xbb\xbf")
    assert job.row_count_actual == 2
    assert job.file_size_bytes == len(content)
    assert job.content_type == "text/csv; charset=utf-8"
    assert job.started_at is not None
    assert job.finished_at is not None
    assert job.expires_at == pytest.approx(
        job.finished_at + dt.timedelta(days=settings.EXPORT_JOB_RETENTION_DAYS),
        rel=0,
    )


@pytest.mark.django_db
def test_process_export_jobs_completes_queued_xlsx_job(tmp_path, settings, users, schema, records):
    settings.MEDIA_ROOT = tmp_path
    job = create_job(
        schema,
        users["viewer"],
        "format=xlsx&at=2024-06-30&search=A-001",
    )

    summary = process_export_jobs(limit=1)

    job.refresh_from_db()
    workbook = load_workbook(BytesIO(job.file.read()), data_only=True)
    metadata = {
        workbook["metadata"].cell(row=index, column=1).value: workbook["metadata"].cell(
            row=index,
            column=2,
        ).value
        for index in range(1, workbook["metadata"].max_row + 1)
    }

    assert summary["processed"] == 1
    assert job.status == ExportJob.Status.COMPLETED
    assert job.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert workbook.sheetnames == ["data", "metadata"]
    assert workbook["data"].max_row == 2
    assert metadata["row_count"] == 1
    assert metadata["format"] == "xlsx"
    snapshot = json.loads(metadata["query_snapshot"])
    assert snapshot["requested_at"] == job.query_snapshot["requested_at"]
    assert snapshot["row_count"] == 1


@pytest.mark.django_db
def test_process_export_jobs_uses_query_snapshot_instead_of_mutated_job_fields(
    tmp_path, settings, users, schema, records
):
    settings.MEDIA_ROOT = tmp_path
    job = create_job(schema, users["viewer"], "format=csv&at=2024-06-30&ordering=-business_code")
    job.export_format = ExportJob.Format.XLSX
    job.save(update_fields=["export_format"])

    process_export_jobs(limit=1)

    job.refresh_from_db()
    lines = job.file.read().decode("utf-8-sig").splitlines()

    assert job.status == ExportJob.Status.COMPLETED
    assert job.export_format == ExportJob.Format.CSV
    assert job.filename.endswith(".csv")
    assert job.content_type == "text/csv; charset=utf-8"
    assert lines[1].startswith("B-001")


@pytest.mark.django_db
def test_process_export_jobs_uses_snapshot_schema_version_fields(
    tmp_path, settings, users, schema, records
):
    settings.MEDIA_ROOT = tmp_path
    job = create_job(schema, users["viewer"], "format=csv&at=2024-06-30")
    schema.fields_config = [
        *schema.fields_config,
        {
            "key": "serial_no",
            "label": "Serial No",
            "type": "text",
            "introduced_in_version": 2,
        },
    ]
    schema.current_version = 2
    schema.save(update_fields=["fields_config", "current_version"])
    SchemaVersion.objects.create(
        schema=schema,
        version=2,
        fields_config=schema.fields_config,
        changelog="Add serial number",
        created_by=users["owner"],
    )

    process_export_jobs(limit=1)

    job.refresh_from_db()
    header = job.file.read().decode("utf-8-sig").splitlines()[0]

    assert job.status == ExportJob.Status.COMPLETED
    assert "Serial No" not in header
    assert "Asset No,Status,Owner" in header


@pytest.mark.django_db
def test_process_export_jobs_uses_selected_export_columns_for_csv(tmp_path, settings, users, schema, records):
    settings.MEDIA_ROOT = tmp_path
    schema.fields_config = [
        *schema.fields_config,
        {"key": "serial_no", "label": "Serial No", "type": "text", "introduced_in_version": 1},
    ]
    schema.save(update_fields=["fields_config"])
    SchemaVersion.objects.filter(schema=schema, version=1).update(fields_config=schema.fields_config)
    TemporalRecord.objects.filter(entity__schema=schema).update(
        data_payload={"asset_no": "A-001", "status": "In Use", "owner": "Alice", "serial_no": "SN-1"}
    )
    job = create_job_from_spec(
        schema,
        users["viewer"],
        export_spec(
            schema,
            export_format="csv",
            columns={"mode": "selected", "field_keys": ["owner", "asset_no"]},
        ),
    )

    process_export_jobs(limit=1)

    job.refresh_from_db()
    header = job.file.read().decode("utf-8-sig").splitlines()[0]

    assert job.status == ExportJob.Status.COMPLETED
    assert header == "display_code,valid_from,valid_to,schema_version,Owner,Asset No"
    assert "Status" not in header
    assert "Serial No" not in header


@pytest.mark.django_db
def test_process_export_jobs_uses_selected_export_columns_for_xlsx_metadata(
    tmp_path, settings, users, schema, records
):
    settings.MEDIA_ROOT = tmp_path
    job = create_job_from_spec(
        schema,
        users["viewer"],
        export_spec(
            schema,
            export_format="xlsx",
            columns={"mode": "selected", "field_keys": ["status"]},
        ),
    )

    process_export_jobs(limit=1)

    job.refresh_from_db()
    workbook = load_workbook(BytesIO(job.file.read()), data_only=True)
    data_sheet = workbook["data"]
    headers = [data_sheet.cell(row=1, column=column).value for column in range(1, data_sheet.max_column + 1)]
    metadata = {
        workbook["metadata"].cell(row=index, column=1).value: workbook["metadata"].cell(row=index, column=2).value
        for index in range(1, workbook["metadata"].max_row + 1)
    }

    assert headers == ["display_code", "valid_from", "valid_to", "schema_version", "Status"]
    assert metadata["export_column_mode"] == "selected"
    assert metadata["export_column_count"] == 1
    assert json.loads(metadata["export_column_keys"]) == ["status"]


@pytest.mark.django_db
def test_process_export_jobs_excludes_unviewable_selected_fields(tmp_path, settings, users, schema, records):
    settings.MEDIA_ROOT = tmp_path
    schema.fields_config = [
        *schema.fields_config,
        {
            "key": "secret",
            "label": "Secret",
            "type": "text",
            "sensitive": True,
            "masking": {"visible_roles": ["owner"]},
            "introduced_in_version": 1,
        },
    ]
    schema.save(update_fields=["fields_config"])
    SchemaVersion.objects.filter(schema=schema, version=1).update(fields_config=schema.fields_config)
    job = create_job_from_spec(
        schema,
        users["viewer"],
        export_spec(
            schema,
            export_format="csv",
            columns={"mode": "selected", "field_keys": ["asset_no", "secret"]},
        ),
    )

    process_export_jobs(limit=1)

    job.refresh_from_db()
    text = job.file.read().decode("utf-8-sig")
    header = text.splitlines()[0]

    assert "Asset No" in header
    assert "Secret" not in header
    assert "***" not in text


@pytest.mark.django_db
def test_process_export_jobs_replays_current_page_row_scope_entity_ids(
    tmp_path, settings, users, schema, records
):
    settings.MEDIA_ROOT = tmp_path
    asset_b = Entity.objects.get(schema=schema, business_code="B-001")
    job = create_job_from_spec(
        schema,
        users["viewer"],
        export_spec(
            schema,
            export_format="csv",
            row_scope={"mode": "current_page", "selected_entity_ids": [asset_b.id]},
        ),
    )

    process_export_jobs(limit=1)

    job.refresh_from_db()
    lines = job.file.read().decode("utf-8-sig").splitlines()

    assert job.status == ExportJob.Status.COMPLETED
    assert job.row_count_actual == 1
    assert len(lines) == 2
    assert lines[1].startswith("B-001")
    assert "A-001" not in "\n".join(lines[1:])


@pytest.mark.django_db
def test_process_export_jobs_replays_selected_entities_row_scope_order(
    tmp_path, settings, users, schema, records
):
    settings.MEDIA_ROOT = tmp_path
    asset_a = Entity.objects.get(schema=schema, business_code="A-001")
    asset_b = Entity.objects.get(schema=schema, business_code="B-001")
    job = create_job_from_spec(
        schema,
        users["viewer"],
        export_spec(
            schema,
            export_format="csv",
            row_scope={"mode": "selected_entities", "selected_entity_ids": [asset_b.id, asset_a.id]},
            ordering="business_code",
        ),
    )

    process_export_jobs(limit=1)

    job.refresh_from_db()
    lines = job.file.read().decode("utf-8-sig").splitlines()

    assert job.status == ExportJob.Status.COMPLETED
    assert job.row_count_actual == 2
    assert lines[1].startswith("B-001")
    assert lines[2].startswith("A-001")


@pytest.mark.django_db
def test_process_export_jobs_writes_row_scope_metadata(tmp_path, settings, users, schema, records):
    settings.MEDIA_ROOT = tmp_path
    asset_b = Entity.objects.get(schema=schema, business_code="B-001")
    job = create_job_from_spec(
        schema,
        users["viewer"],
        export_spec(
            schema,
            export_format="xlsx",
            row_scope={"mode": "selected_entities", "selected_entity_ids": [asset_b.id]},
        ),
    )

    process_export_jobs(limit=1)

    job.refresh_from_db()
    workbook = load_workbook(BytesIO(job.file.read()), data_only=True)
    metadata = {
        workbook["metadata"].cell(row=index, column=1).value: workbook["metadata"].cell(
            row=index,
            column=2,
        ).value
        for index in range(1, workbook["metadata"].max_row + 1)
    }

    assert metadata["export_row_scope_mode"] == "selected_entities"
    assert metadata["export_row_scope_selected_entity_count"] == 1


@pytest.mark.django_db
def test_process_export_jobs_uses_snapshot_schema_version_for_ordering(
    tmp_path, settings, users, schema, records
):
    settings.MEDIA_ROOT = tmp_path
    record_a = TemporalRecord.objects.get(data_payload__asset_no="A-001", valid_to__isnull=True)
    record_b = TemporalRecord.objects.get(data_payload__asset_no="B-001")
    record_a.data_payload = {"asset_no": "A-001", "status": "Repair", "owner": "10"}
    record_b.data_payload = {"asset_no": "B-001", "status": "In Use", "owner": "2"}
    record_a.save(update_fields=["data_payload"])
    record_b.save(update_fields=["data_payload"])
    job = create_job(schema, users["viewer"], "format=csv&at=2024-06-30&ordering=owner")
    schema.fields_config = [
        field | {"type": "number"} if field["key"] == "owner" else field
        for field in schema.fields_config
    ]
    schema.current_version = 2
    schema.save(update_fields=["fields_config", "current_version"])
    SchemaVersion.objects.create(
        schema=schema,
        version=2,
        fields_config=schema.fields_config,
        changelog="Treat owner as a number",
        created_by=users["owner"],
    )

    process_export_jobs(limit=1)

    job.refresh_from_db()
    lines = job.file.read().decode("utf-8-sig").splitlines()

    assert job.status == ExportJob.Status.COMPLETED
    assert lines[1].startswith("A-001")
    assert lines[2].startswith("B-001")


@pytest.mark.django_db
def test_process_export_jobs_fails_when_snapshot_schema_version_is_missing(
    tmp_path, settings, users, schema, records
):
    settings.MEDIA_ROOT = tmp_path
    job = create_job(schema, users["viewer"], "format=csv&at=2024-06-30")
    SchemaVersion.objects.filter(schema=schema, version=job.query_snapshot["schema_version"]).delete()

    summary = process_export_jobs(limit=1)

    job.refresh_from_db()

    assert summary["failed"] == 1
    assert job.status == ExportJob.Status.FAILED
    assert job.error_code == "schema_version_missing"
    assert job.file.name == ""


@pytest.mark.django_db
def test_process_export_jobs_marks_job_failed_when_export_generation_raises(
    tmp_path, settings, users, schema, records, monkeypatch
):
    settings.MEDIA_ROOT = tmp_path
    job = create_job(schema, users["viewer"], "format=csv&at=2024-06-30")

    def raise_error(*args, **kwargs):
        raise RuntimeError("boom")

    monkeypatch.setattr(export_job_worker, "build_current_export", raise_error)

    summary = process_export_jobs(limit=1)

    job.refresh_from_db()

    assert summary["failed"] == 1
    assert job.status == ExportJob.Status.FAILED
    assert job.error_code == "export_generation_failed"
    assert job.error_message == "boom"
    assert "Traceback" not in job.error_message
    assert job.finished_at is not None
    assert job.file.name == ""


@pytest.mark.django_db
def test_process_export_jobs_marks_stale_running_job_failed(tmp_path, settings, users, schema):
    settings.MEDIA_ROOT = tmp_path
    now = timezone.make_aware(dt.datetime(2026, 5, 25, 21, 0, 0))
    job = create_job(schema, users["viewer"], "format=csv&at=2024-06-30")
    job.status = ExportJob.Status.RUNNING
    job.started_at = now - dt.timedelta(minutes=31)
    job.save(update_fields=["status", "started_at"])

    summary = process_export_jobs(limit=0, now=now)

    job.refresh_from_db()

    assert summary["stale_failed"] == 1
    assert job.status == ExportJob.Status.FAILED
    assert job.error_code == "worker_timeout"
    assert job.finished_at == now


@pytest.mark.django_db
def test_process_export_jobs_cleanup_expired_files(tmp_path, settings, users, schema):
    settings.MEDIA_ROOT = tmp_path
    now = timezone.make_aware(dt.datetime(2026, 5, 25, 21, 0, 0))
    job = create_job(schema, users["viewer"], "format=csv&at=2024-06-30")
    job.status = ExportJob.Status.COMPLETED
    job.finished_at = now - dt.timedelta(days=31)
    job.expires_at = now - dt.timedelta(minutes=1)
    job.file.save("expired.csv", ContentFile(b"\xef\xbb\xbfheader\nrow"), save=False)
    expired_path = job.file.path
    job.filename = "expired.csv"
    job.content_type = "text/csv; charset=utf-8"
    job.file_size_bytes = 13
    job.save(
        update_fields=[
            "status",
            "finished_at",
            "expires_at",
            "file",
            "filename",
            "content_type",
            "file_size_bytes",
        ]
    )

    summary = process_export_jobs(limit=0, now=now, cleanup_expired=True)

    job.refresh_from_db()

    assert summary["expired"] == 1
    assert job.status == ExportJob.Status.EXPIRED
    assert job.file.name == ""
    assert not Path(expired_path).exists()


@pytest.mark.django_db
def test_process_export_jobs_command_processes_once_and_cleans_up(tmp_path, settings, users, schema):
    settings.MEDIA_ROOT = tmp_path
    now = timezone.make_aware(dt.datetime(2026, 5, 25, 21, 0, 0))
    queued = create_job(schema, users["viewer"], "format=csv&at=2024-06-30")
    expired = create_job(schema, users["viewer"], "format=csv&at=2024-06-29")
    expired.status = ExportJob.Status.COMPLETED
    expired.finished_at = now - dt.timedelta(days=31)
    expired.expires_at = now - dt.timedelta(minutes=1)
    expired.file.save("expired.csv", ContentFile(b"\xef\xbb\xbfheader\nrow"), save=False)
    expired.save(update_fields=["status", "finished_at", "expires_at", "file"])

    out = StringIO()
    call_command(
        "process_export_jobs",
        "--once",
        "--limit=1",
        "--cleanup-expired",
        stdout=out,
    )

    queued.refresh_from_db()
    expired.refresh_from_db()

    assert queued.status == ExportJob.Status.COMPLETED
    assert expired.status == ExportJob.Status.EXPIRED
    assert "processed=1" in out.getvalue()


@pytest.mark.django_db
def test_process_export_jobs_command_supports_loop_and_sleep(monkeypatch):
    calls = []

    def fake_process_export_jobs(**kwargs):
        calls.append(kwargs)
        return {"processed": 0, "failed": 0, "stale_failed": 0, "expired": 0}

    def stop_sleep(seconds):
        assert seconds == 0.0
        raise KeyboardInterrupt

    monkeypatch.setattr(process_export_jobs_command, "process_export_jobs", fake_process_export_jobs)
    monkeypatch.setattr(process_export_jobs_command.time, "sleep", stop_sleep)

    out = StringIO()
    call_command("process_export_jobs", "--loop", "--sleep=0", "--limit=2", stdout=out)

    assert calls == [{"limit": 2, "cleanup_expired": False}]
    assert "stopped" in out.getvalue()


def export_spec(
    schema,
    *,
    export_format: str = "xlsx",
    at: str = "2024-06-30",
    retro: bool = False,
    row_scope: dict | None = None,
    filters: list[dict] | None = None,
    search: str = "",
    ordering: str = "business_code",
    change_set: int | None = None,
    columns: dict | None = None,
    schema_version: int = 1,
) -> dict:
    return {
        "schema_id": schema.id,
        "schema_version": schema_version,
        "scope": "current_view",
        "format": export_format,
        "time": {"at": at, "retro": retro},
        "row_scope": row_scope or {"mode": "filtered_result", "selected_entity_ids": []},
        "filters": filters or [],
        "search": search,
        "ordering": ordering,
        "change_set": change_set,
        "columns": columns or {"mode": "all_exportable", "field_keys": []},
    }

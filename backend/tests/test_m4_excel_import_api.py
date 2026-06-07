import datetime as dt
import json
from io import BytesIO

import pytest
from django.contrib.auth.models import User
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIClient

from apps.changesets.models import ChangeSet
from apps.schemas.models import DataSchema, TableCollaborator
from apps.temporal.models import Entity, TemporalRecord


@pytest.fixture
def users(db):
    return {
        "owner": User.objects.create_user(username="owner", password="pass"),
        "editor": User.objects.create_user(username="editor", password="pass"),
        "viewer": User.objects.create_user(username="viewer", password="pass"),
    }


@pytest.fixture
def client():
    return APIClient()


def auth(client, user):
    client.force_authenticate(user=user)
    return client


@pytest.fixture
def schema(users):
    schema = DataSchema.objects.create(
        schema_code="asset_list",
        name="固定资产表",
        temporal_mode="continuous",
        identity_field_key="asset_no",
        fields_config=[
            {"key": "asset_no", "label": "资产编号", "type": "text", "required": True},
            {
                "key": "status",
                "label": "状态",
                "type": "enum",
                "validators": {"options": ["在用", "维修", "报废"]},
            },
            {"key": "owner", "label": "负责人", "type": "text"},
        ],
        current_version=1,
        owner=users["owner"],
        visibility="shared",
        created_by=users["owner"],
    )
    TableCollaborator.objects.create(
        schema=schema,
        user=users["editor"],
        role="editor",
        added_by=users["owner"],
    )
    TableCollaborator.objects.create(
        schema=schema,
        user=users["viewer"],
        role="viewer",
        added_by=users["owner"],
    )
    return schema


@pytest.fixture
def records(schema, users):
    change_set = ChangeSet.objects.create(
        schema=schema,
        summary="初始数据",
        status=ChangeSet.Status.APPLIED,
        created_by=users["owner"],
        applied_at=timezone.now(),
    )
    asset_a = Entity.objects.create(schema=schema, business_code="A-001", created_by=users["owner"])
    asset_b = Entity.objects.create(schema=schema, business_code="B-001", created_by=users["owner"])
    TemporalRecord.objects.create(
        entity=asset_a,
        schema_version=1,
        data_payload={"asset_no": "A-001", "status": "维修", "owner": "张三"},
        valid_from=dt.date(2024, 6, 1),
        change_set=change_set,
        recorded_by=users["owner"],
    )
    TemporalRecord.objects.create(
        entity=asset_b,
        schema_version=1,
        data_payload={"asset_no": "B-001", "status": "在用", "owner": "李四"},
        valid_from=dt.date(2024, 3, 1),
        change_set=change_set,
        recorded_by=users["owner"],
    )
    return {"asset_a": asset_a, "asset_b": asset_b}


@pytest.mark.django_db
def test_import_template_download_contains_headers_and_comments(client, users, schema):
    response = auth(client, users["viewer"]).get(f"/api/v1/schemas/{schema.id}/import/template")

    assert response.status_code == 200
    assert response["Content-Type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active

    assert [cell.value for cell in sheet[1]][:4] == ["资产编号", "状态", "负责人", "valid_from"]
    assert sheet["A1"].comment is not None
    assert "asset_no" in sheet["A1"].comment.text
    assert "在用" in sheet["B1"].comment.text


@pytest.mark.django_db
def test_import_preview_classifies_create_update_missing_and_invalid(
    client, users, schema, records
):
    workbook = make_workbook(
        [
            ["资产编号", "状态", "负责人", "valid_from"],
            ["A-001", "报废", "张三", "2024-08-01"],
            ["C-001", "在用", "王五", "2024-08-01"],
            ["D-001", "非法状态", "赵六", "2024-08-01"],
        ]
    )

    response = auth(client, users["editor"]).post(
        f"/api/v1/schemas/{schema.id}/import/preview",
        {"file": workbook, "at": "2024-08-01", "missing_policy": "terminate"},
        format="multipart",
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["summary"] == {
        "create": 1,
        "update": 1,
        "missing": 1,
        "invalid": 1,
        "unchanged": 0,
    }
    assert [row["action"] for row in payload["rows"]] == ["update", "create", "invalid"]
    assert [row["display_code"] for row in payload["rows"]] == ["A-001", "C-001", "D-001"]
    assert payload["rows"][0]["changed_fields"] == ["status"]
    assert payload["rows"][2]["errors"][0]["path"] == "data_payload.status"
    assert payload["missing"][0]["business_code"] == "B-001"
    assert payload["missing"][0]["display_code"] == "B-001"


@pytest.mark.django_db
def test_import_preview_reports_duplicate_identity_diagnostics(client, users, schema, records):
    workbook = make_workbook(
        [
            ["资产编号", "状态", "负责人", "valid_from"],
            ["A-001", "报废", "张三", "2024-08-01"],
            ["A-001", "维修", "李四", "2024-08-01"],
            ["C-001", "在用", "王五", "2024-08-01"],
        ]
    )

    response = auth(client, users["editor"]).post(
        f"/api/v1/schemas/{schema.id}/import/preview",
        {"file": workbook, "at": "2024-08-01", "missing_policy": "terminate"},
        format="multipart",
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["summary"] == {
        "create": 1,
        "update": 0,
        "missing": 1,
        "invalid": 2,
        "unchanged": 0,
    }
    assert payload["identity_diagnostics"] == {
        "mode": "single",
        "status": "error",
        "identity_field_key": "asset_no",
        "identity_field_label": "资产编号",
        "message": "当前实体标识字段“资产编号”存在重复值，资产编号不适合作为实体标识。请选择员工号、证件号派生码、社保账号，或创建组合标识字段。",
        "duplicate_values": [{"value": "A-001", "count": 2, "row_numbers": [2, 3]}],
    }
    duplicate_rows = [row for row in payload["rows"] if row["business_code"] == "A-001"]
    assert len(duplicate_rows) == 2
    assert all(row["action"] == "invalid" for row in duplicate_rows)
    assert all(
        any(error["code"] == "duplicate_identity" for error in row["errors"])
        for row in duplicate_rows
    )


@pytest.mark.django_db
def test_import_commit_creates_draft_changeset_entries(client, users, schema, records):
    workbook = make_workbook(
        [
            ["资产编号", "状态", "负责人", "valid_from"],
            ["A-001", "报废", "张三", "2024-08-01"],
            ["C-001", "在用", "王五", "2024-08-01"],
        ]
    )

    response = auth(client, users["editor"]).post(
        f"/api/v1/schemas/{schema.id}/import/commit",
        {
            "file": workbook,
            "at": "2024-08-01",
            "missing_policy": "terminate",
            "summary": "Excel 资产盘点",
        },
        format="multipart",
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["status"] == "draft"
    assert payload["source"] == "excel"
    assert payload["summary"] == "Excel 资产盘点"
    assert payload["action_counts"] == {"create": 1, "update": 1, "terminate": 1}
    assert [entry["action"] for entry in payload["entries"]] == ["update", "terminate", "create"]


@pytest.mark.django_db
def test_import_preview_supports_manual_mapping_and_commit_corrected_rows(
    client, users, schema, records
):
    workbook = make_workbook(
        [
            ["编号", "状态", "负责人", "valid_from"],
            ["D-001", "非法状态", "赵六", "2024-08-01"],
        ]
    )

    preview = auth(client, users["editor"]).post(
        f"/api/v1/schemas/{schema.id}/import/preview",
        {
            "file": workbook,
            "at": "2024-08-01",
            "mappings_json": json.dumps([{"source_column": "编号", "field_key": "asset_no"}]),
        },
        format="multipart",
    )

    assert preview.status_code == 200
    preview_payload = preview.json()
    assert preview_payload["mappings"][0]["field_key"] == "asset_no"
    assert preview_payload["summary"]["invalid"] == 1

    corrected = preview_payload["rows"][0]
    corrected["action"] = "create"
    corrected["data_after"]["status"] = "在用"
    corrected["errors"] = []
    commit = client.post(
        f"/api/v1/schemas/{schema.id}/import/commit",
        {
            "summary": "修正后导入",
            "rows_json": json.dumps([corrected]),
            "missing_json": json.dumps([]),
        },
        format="json",
    )

    assert commit.status_code == 201
    assert commit.json()["action_counts"] == {"create": 1, "update": 0, "terminate": 0}


@pytest.mark.django_db
def test_import_preview_reports_invalid_person_values(client, users, schema):
    schema.fields_config = [
        *schema.fields_config,
        {"key": "assignee", "label": "Assignee", "type": "person"},
    ]
    schema.save(update_fields=["fields_config"])
    workbook = make_workbook(
        [
            ["asset_no", "assignee"],
            ["P-001", "John"],
        ]
    )

    response = auth(client, users["editor"]).post(
        f"/api/v1/schemas/{schema.id}/import/preview",
        {"file": workbook, "at": "2024-08-01"},
        format="multipart",
    )

    assert response.status_code == 400
    assert "assignee" in response.json()


@pytest.mark.django_db
def test_import_preview_rejects_unknown_mapping_field_key(client, users, schema):
    workbook = make_workbook(
        [
            ["asset_no"],
            ["M-001"],
        ]
    )

    response = auth(client, users["editor"]).post(
        f"/api/v1/schemas/{schema.id}/import/preview",
        {
            "file": workbook,
            "at": "2024-08-01",
            "mappings_json": json.dumps(
                [{"source_column": "asset_no", "field_key": "missing_field"}]
            ),
        },
        format="multipart",
    )

    assert response.status_code == 400
    assert "mappings_json" in response.json()


def make_workbook(rows: list[list[object]]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    buffer.name = "assets.xlsx"
    return buffer

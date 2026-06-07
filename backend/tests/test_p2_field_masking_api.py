import datetime as dt
from io import BytesIO

import pytest
from django.contrib.auth.models import User
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.changesets.models import ChangeEntry, ChangeSet
from apps.schemas.models import DataSchema, TableCollaborator
from apps.temporal.models import Entity, TemporalRecord


@pytest.fixture
def users(db):
    return {
        "owner": User.objects.create_user(username="owner", password="pass"),
        "viewer": User.objects.create_user(username="viewer", password="pass"),
    }


@pytest.fixture
def client():
    return APIClient()


def auth(client, user):
    client.force_authenticate(user=user)
    return client


@pytest.fixture
def masked_schema(users):
    schema = DataSchema.objects.create(
        schema_code="employee_cards",
        name="Employee Cards",
        temporal_mode="continuous",
        identity_field_key="employee_no",
        fields_config=[
            {"key": "employee_no", "label": "Employee No", "type": "text", "required": True},
            {
                "key": "id_card",
                "label": "ID Card",
                "type": "text",
                "sensitive": True,
                "masking": {"mode": "partial", "visible_roles": ["admin", "owner"]},
            },
            {
                "key": "salary",
                "label": "Salary",
                "type": "number",
                "sensitive": True,
                "masking": {"mode": "full", "visible_roles": ["owner"]},
            },
        ],
        owner=users["owner"],
        visibility="shared",
        created_by=users["owner"],
    )
    TableCollaborator.objects.create(
        schema=schema,
        user=users["viewer"],
        role="viewer",
        added_by=users["owner"],
    )
    return schema


@pytest.fixture
def masked_records(masked_schema, users):
    change_set = ChangeSet.objects.create(
        schema=masked_schema,
        summary="seed",
        status=ChangeSet.Status.APPLIED,
        created_by=users["owner"],
        applied_at=timezone.now(),
    )
    entity = Entity.objects.create(
        schema=masked_schema,
        business_code="E-001",
        created_by=users["owner"],
    )
    record = TemporalRecord.objects.create(
        entity=entity,
        schema_version=1,
        data_payload={
            "employee_no": "E-001",
            "id_card": "110105199001011234",
            "salary": 120000,
        },
        valid_from=dt.date(2024, 1, 1),
        change_set=change_set,
        recorded_by=users["owner"],
    )
    ChangeEntry.objects.create(
        change_set=change_set,
        entity=entity,
        action=ChangeEntry.Action.CREATE,
        data_after=record.data_payload,
        valid_from=record.valid_from,
        new_record=record,
    )
    return {"change_set": change_set, "entity": entity, "record": record}


@pytest.mark.django_db
def test_records_mask_sensitive_field_for_viewer(masked_schema, masked_records, users, client):
    owner_response = auth(client, users["owner"]).get(
        f"/api/v1/schemas/{masked_schema.id}/records/?at=2024-01-02"
    )
    viewer_response = auth(client, users["viewer"]).get(
        f"/api/v1/schemas/{masked_schema.id}/records/?at=2024-01-02"
    )

    assert owner_response.status_code == 200
    assert owner_response.json()["results"][0]["data_payload"]["id_card"] == "110105199001011234"
    assert viewer_response.status_code == 200
    payload = viewer_response.json()["results"][0]["data_payload"]
    assert payload["id_card"] == {"kind": "masked", "display": "110***********1234"}
    assert payload["salary"] == {"kind": "masked", "display": "***"}


@pytest.mark.django_db
def test_records_search_and_ordering_do_not_leak_masked_values(
    masked_schema,
    masked_records,
    users,
    client,
):
    search_response = auth(client, users["viewer"]).get(
        f"/api/v1/schemas/{masked_schema.id}/records/?at=2024-01-02&search=110105"
    )
    ordering_response = client.get(
        f"/api/v1/schemas/{masked_schema.id}/records/?at=2024-01-02&ordering=id_card"
    )

    assert search_response.status_code == 200
    assert search_response.json()["count"] == 0
    assert ordering_response.status_code == 400
    assert "ordering" in ordering_response.json()


@pytest.mark.django_db
def test_changeset_detail_masks_sensitive_fields_for_viewer(
    masked_schema,
    masked_records,
    users,
    client,
):
    response = auth(client, users["viewer"]).get(
        f"/api/v1/schemas/{masked_schema.id}/changesets/{masked_records['change_set'].id}/"
    )

    assert response.status_code == 200
    entry = response.json()["entries"][0]
    assert entry["data_after"]["id_card"] == {"kind": "masked", "display": "110***********1234"}
    assert entry["data_after"]["salary"] == {"kind": "masked", "display": "***"}


@pytest.mark.django_db
def test_exports_mask_sensitive_fields_for_viewer(
    masked_schema,
    masked_records,
    users,
    client,
):
    api = auth(client, users["viewer"])

    current_response = api.get(
        f"/api/v1/schemas/{masked_schema.id}/export/current",
        {"format": "csv", "at": "2024-01-02"},
    )
    entity_response = api.get(f"/api/v1/entities/{masked_records['entity'].id}/export")
    changeset_response = api.get(
        f"/api/v1/changesets/{masked_records['change_set'].id}/export"
    )

    assert current_response.status_code == 200, current_response.content.decode()
    current_text = current_response.content.decode("utf-8-sig")
    assert "110105199001011234" not in current_text
    assert "120000" not in current_text
    assert "110***********1234" in current_text
    assert "***" in current_text

    assert entity_response.status_code == 200
    entity_values = workbook_values(entity_response.content)
    assert "110105199001011234" not in entity_values
    assert "120000" not in entity_values
    assert "110***********1234" in entity_values
    assert "***" in entity_values

    assert changeset_response.status_code == 200
    changeset_values = workbook_values(changeset_response.content)
    assert "110105199001011234" not in changeset_values
    assert "120000" not in changeset_values
    assert "110***********1234" in changeset_values
    assert "***" in changeset_values


def workbook_values(content: bytes) -> str:
    workbook = load_workbook(BytesIO(content), data_only=True)
    values = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            values.extend("" if value is None else str(value) for value in row)
    return "\n".join(values)

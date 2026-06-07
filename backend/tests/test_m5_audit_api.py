import csv
from io import BytesIO, StringIO

import pytest
from django.contrib.auth.models import User
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.audit.models import AuditLog
from apps.audit.services import record_audit_log
from apps.changesets.models import ChangeSet
from apps.schemas.models import DataSchema


@pytest.fixture
def users(db):
    return {
        "owner": User.objects.create_user(username="owner", password="pass"),
        "editor": User.objects.create_user(username="editor", password="pass"),
        "viewer": User.objects.create_user(username="viewer", password="pass"),
        "outsider": User.objects.create_user(username="outsider", password="pass"),
        "admin": User.objects.create_superuser(
            username="admin",
            email="admin@example.com",
            password="pass",
        ),
    }


@pytest.fixture
def client():
    return APIClient()


def auth(client, user):
    client.force_authenticate(user=user)
    return client


def make_schema(schema_code, owner, visibility="shared"):
    return DataSchema.objects.create(
        schema_code=schema_code,
        name=schema_code.replace("_", " ").title(),
        description="内部资产台账",
        icon="box",
        temporal_mode="continuous",
        identity_field_key="asset_no",
        fields_config=[{"key": "asset_no", "label": "资产编号", "type": "text"}],
        owner=owner,
        visibility=visibility,
        created_by=owner,
    )


def make_changeset(schema, user):
    return ChangeSet.objects.create(
        schema=schema,
        summary=f"{schema.schema_code} 变更",
        status=ChangeSet.Status.APPLIED,
        created_by=user,
    )


def result_actions(response):
    return [item["action"] for item in response.json()["results"]]


@pytest.mark.django_db
def test_audit_logs_are_filtered_for_admin_owner_and_actor(client, users):
    owned_schema = make_schema("owned_assets", users["owner"])
    other_schema = make_schema("other_assets", users["outsider"])
    owned_changeset = make_changeset(owned_schema, users["editor"])

    record_audit_log(
        actor=users["editor"],
        action="collaborator.add",
        target_type="schema",
        target_id=owned_schema.id,
        detail={"user_id": users["viewer"].id},
    )
    record_audit_log(
        actor=users["editor"],
        action="changeset.revert",
        target_type="changeset",
        target_id=owned_changeset.id,
        detail={"revert_changeset_id": 999},
    )
    record_audit_log(
        actor=users["viewer"],
        action="data.export",
        target_type="schema",
        target_id=owned_schema.id,
        detail={"row_count": 20},
    )
    record_audit_log(
        actor=users["outsider"],
        action="schema.archive",
        target_type="schema",
        target_id=other_schema.id,
        detail={"schema_code": other_schema.schema_code},
    )

    admin_response = auth(client, users["admin"]).get("/api/v1/audit-logs/")
    owner_response = auth(client, users["owner"]).get("/api/v1/audit-logs/")
    viewer_response = auth(client, users["viewer"]).get("/api/v1/audit-logs/")

    assert admin_response.status_code == 200
    assert owner_response.status_code == 200
    assert viewer_response.status_code == 200
    assert admin_response.json()["count"] == 4
    assert result_actions(owner_response) == [
        "data.export",
        "changeset.revert",
        "collaborator.add",
    ]
    assert result_actions(viewer_response) == ["data.export"]


@pytest.mark.django_db
def test_audit_logs_support_actor_target_time_and_sensitive_filters(client, users):
    schema = make_schema("asset_list", users["owner"])
    record_audit_log(
        actor=users["owner"],
        action="schema.visibility_change",
        target_type="schema",
        target_id=schema.id,
        detail={"from_visibility": "shared", "to_visibility": "public"},
    )
    record_audit_log(
        actor=users["editor"],
        action="collaborator.add",
        target_type="schema",
        target_id=schema.id,
        detail={"user_id": users["viewer"].id},
    )

    response = auth(client, users["admin"]).get(
        "/api/v1/audit-logs/",
        {
            "actor_id": users["owner"].id,
            "target_type": "schema",
            "target_id": schema.id,
            "created_after": "2999-01-01",
        },
    )
    assert response.status_code == 200
    assert response.json()["count"] == 0

    response = client.get(
        "/api/v1/audit-logs/",
        {
            "actor": "own",
            "target_type": "schema",
            "target_id": schema.id,
            "is_sensitive": "true",
            "page_size": 1,
        },
    )

    assert response.status_code == 200
    assert response.json()["count"] == 1
    assert response.json()["page_size"] == 1
    assert response.json()["results"][0]["action"] == "schema.visibility_change"
    assert response.json()["results"][0]["target_schema_id"] == schema.id
    assert response.json()["results"][0]["target_schema_name"] == schema.name


@pytest.mark.django_db
def test_sensitive_audit_logs_are_admin_only(client, users):
    schema = make_schema("asset_list", users["owner"])
    record_audit_log(
        actor=users["owner"],
        action="schema.visibility_change",
        target_type="schema",
        target_id=schema.id,
        detail={"from_visibility": "private", "to_visibility": "public"},
    )
    record_audit_log(
        actor=users["owner"],
        action="collaborator.add",
        target_type="schema",
        target_id=schema.id,
        detail={"user_id": users["viewer"].id},
    )

    denied = auth(client, users["owner"]).get("/api/v1/audit-logs/sensitive")
    allowed = auth(client, users["admin"]).get("/api/v1/audit-logs/sensitive")

    assert denied.status_code == 403
    assert allowed.status_code == 200
    assert allowed.json()["count"] == 1
    assert allowed.json()["results"][0]["is_sensitive"] is True
    assert allowed.json()["results"][0]["action"] == "schema.visibility_change"
    assert AuditLog.objects.count() == 2


@pytest.mark.django_db
def test_sensitive_audit_export_csv_respects_filters_and_records_audit(client, users):
    schema = make_schema("asset_list", users["owner"])
    record_audit_log(
        actor=users["owner"],
        action="schema.visibility_change",
        target_type="schema",
        target_id=schema.id,
        detail={"from_visibility": "private", "to_visibility": "public"},
    )
    record_audit_log(
        actor=users["editor"],
        action="admin.password_reset",
        target_type="user",
        target_id=users["viewer"].id,
        detail={"username": users["viewer"].username},
    )
    response = auth(client, users["admin"]).get(
        "/api/v1/audit-logs/sensitive/export",
        {"format": "csv", "action": "schema.visibility_change"},
    )

    assert response.status_code == 200
    assert response["Content-Type"].startswith("text/csv")
    rows = list(csv.DictReader(StringIO(response.content.decode("utf-8-sig"))))
    assert [row["action"] for row in rows] == ["schema.visibility_change"]
    assert rows[0]["target_schema_name"] == schema.name
    assert '"to_visibility": "public"' in rows[0]["detail"]
    export_log = AuditLog.objects.filter(action="audit.export").latest("id")
    assert export_log.is_sensitive is True
    assert export_log.detail["format"] == "csv"
    assert export_log.detail["row_count"] == 1
    assert export_log.detail["filters"] == {"action": "schema.visibility_change"}


@pytest.mark.django_db
def test_sensitive_audit_export_xlsx_contains_metadata(client, users):
    schema = make_schema("asset_list", users["owner"])
    record_audit_log(
        actor=users["owner"],
        action="schema.visibility_change",
        target_type="schema",
        target_id=schema.id,
        detail={"from_visibility": "private", "to_visibility": "public"},
    )

    response = auth(client, users["admin"]).get(
        "/api/v1/audit-logs/sensitive/export",
        {"format": "xlsx", "target_type": "schema"},
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook.sheetnames == ["data", "metadata"]
    data_sheet = workbook["data"]
    metadata = {
        workbook["metadata"].cell(row=index, column=1).value: workbook["metadata"].cell(
            row=index, column=2
        ).value
        for index in range(1, workbook["metadata"].max_row + 1)
    }
    assert data_sheet.cell(row=1, column=1).value == "id"
    assert data_sheet.cell(row=2, column=4).value == "schema.visibility_change"
    assert metadata["export_scope"] == "sensitive_audit_logs"
    assert metadata["row_count"] == 1
    assert metadata["format"] == "xlsx"


@pytest.mark.django_db
def test_sensitive_audit_export_is_admin_only(client, users):
    response = auth(client, users["owner"]).get(
        "/api/v1/audit-logs/sensitive/export",
        {"format": "csv"},
    )

    assert response.status_code == 403
    assert not AuditLog.objects.filter(action="audit.export").exists()


@pytest.mark.django_db
def test_admin_password_reset_audit_log_is_sensitive(users):
    log = record_audit_log(
        actor=users["admin"],
        action="admin.password_reset",
        target_type="user",
        target_id=users["viewer"].id,
        detail={"username": users["viewer"].username},
    )

    assert log.is_sensitive is True

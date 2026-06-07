from __future__ import annotations

import datetime as dt
from io import BytesIO

import pytest
from django.contrib.auth.models import User
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIClient

from apps.changesets.models import ChangeEntry, ChangeSet
from apps.schemas.identity import IDENTITY_CODE_FIELD_KEY, build_composite_identity_code
from apps.schemas.models import DataSchema, TableCollaborator
from apps.temporal.models import Entity, TemporalRecord


@pytest.fixture
def users(db):
    return {
        "owner": User.objects.create_user(username="owner", password="pass"),
        "editor": User.objects.create_user(username="editor", password="pass"),
    }


@pytest.fixture
def client():
    return APIClient()


def auth(client, user):
    client.force_authenticate(user=user)
    return client


@pytest.fixture
def composite_schema(users):
    schema = DataSchema.objects.create(
        schema_code="social_base",
        name="社保基数",
        temporal_mode="continuous",
        identity_field_key=IDENTITY_CODE_FIELD_KEY,
        fields_config=composite_fields_config(),
        current_version=1,
        owner=users["owner"],
        visibility="shared",
        created_by=users["owner"],
    )
    TableCollaborator.objects.create(
        schema=schema,
        user=users["editor"],
        role="editor",
        added_by=users["owner"],
    )
    return schema


@pytest.fixture
def composite_record(composite_schema, users):
    change_set = ChangeSet.objects.create(
        schema=composite_schema,
        summary="初始社保数据",
        status=ChangeSet.Status.APPLIED,
        created_by=users["owner"],
        applied_at=timezone.now(),
    )
    entity = Entity.objects.create(
        schema=composite_schema,
        business_code="C01|E001",
        created_by=users["owner"],
    )
    record = TemporalRecord.objects.create(
        entity=entity,
        schema_version=1,
        data_payload={
            "company_code": "C01",
            "employee_no": "E001",
            "name": "张伟",
            "status": "在缴",
            IDENTITY_CODE_FIELD_KEY: "C01|E001",
        },
        valid_from=dt.date(2024, 6, 1),
        change_set=change_set,
        recorded_by=users["owner"],
    )
    ChangeEntry.objects.create(
        change_set=change_set,
        entity=entity,
        action=ChangeEntry.Action.CREATE,
        data_after=record.data_payload,
        valid_from=record.valid_from,
        new_record=record,
    )
    return entity


def test_composite_identity_escapes_separator_and_backslash():
    assert build_composite_identity_code(["C|01", r"E\001"]) == r"C\|01|E\\001"


@pytest.mark.django_db
def test_excel_intake_preview_generates_hidden_composite_identity(client, users):
    token = scan_token(
        client,
        users["owner"],
        workbook_from_rows(
            [
                ["公司编码", "员工号", "姓名", "社保基数"],
                ["C01", "E001", "张伟", 10000],
                ["C02", "E001", "张伟", 12000],
            ]
        ),
    )

    response = client.post(
        "/api/v1/excel-intake/preview",
        {
            "upload_token": token,
            "sheet_name": "Sheet1",
            "header_row": 1,
            "data_start_row": 2,
            "schema": {
                "schema_code": "social_base",
                "name": "社保基数",
                "identity_mode": "composite",
                "identity_field_keys": ["company_code", "employee_no"],
            },
            "fields_config": intake_fields_config(),
            "valid_from": "2026-05-15",
        },
        format="json",
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["schema_draft"]["identity_mode"] == "composite"
    assert payload["schema_draft"]["identity_field_key"] == IDENTITY_CODE_FIELD_KEY
    assert payload["schema_draft"]["identity_field_keys"] == ["company_code", "employee_no"]
    hidden_field = hidden_identity_field(payload["schema_draft"]["fields_config"])
    assert hidden_field["key"] == IDENTITY_CODE_FIELD_KEY
    assert hidden_field["label"] == "实体标识"
    assert hidden_field["type"] == "text"
    assert hidden_field["required"] is True
    assert hidden_field["indexed"] is True
    assert hidden_field["validators"] == {"max_length": 128}
    assert hidden_field["hidden"] is True
    assert hidden_field["system"] is True
    assert hidden_field["identity_mode"] == "composite"
    assert hidden_field["identity_field_keys"] == ["company_code", "employee_no"]
    assert payload["identity_diagnostics"]["mode"] == "composite"
    assert payload["identity_diagnostics"]["identity_field_label"] == "公司编码 + 员工号"
    assert [row["business_code"] for row in payload["rows"]] == ["C01|E001", "C02|E001"]
    assert payload["rows"][0]["data_after"][IDENTITY_CODE_FIELD_KEY] == "C01|E001"
    assert payload["summary"]["invalid"] == 0


@pytest.mark.django_db
def test_excel_intake_preview_marks_duplicate_composite_identity_invalid(client, users):
    token = scan_token(
        client,
        users["owner"],
        workbook_from_rows(
            [
                ["公司编码", "员工号", "姓名"],
                ["C01", "E001", "张伟"],
                ["C01", "E001", "张伟-重复"],
                ["C02", "E001", "李娜"],
            ]
        ),
    )

    response = client.post(
        "/api/v1/excel-intake/preview",
        {
            "upload_token": token,
            "sheet_name": "Sheet1",
            "header_row": 1,
            "data_start_row": 2,
            "schema": {
                "schema_code": "social_base",
                "name": "社保基数",
                "identity_mode": "composite",
                "identity_field_keys": ["company_code", "employee_no"],
            },
            "fields_config": intake_fields_config(include_amount=False),
            "valid_from": "2026-05-15",
        },
        format="json",
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["summary"] == {
        "create": 1,
        "update": 0,
        "missing": 0,
        "invalid": 2,
        "unchanged": 0,
    }
    assert payload["identity_diagnostics"]["status"] == "error"
    assert payload["identity_diagnostics"]["duplicate_values"] == [
        {"value": "C01|E001", "count": 2, "row_numbers": [2, 3]}
    ]
    duplicate_rows = [row for row in payload["rows"] if row["business_code"] == "C01|E001"]
    assert len(duplicate_rows) == 2
    assert all(row["action"] == "invalid" for row in duplicate_rows)
    assert all(
        any(error["code"] == "duplicate_identity" for error in row["errors"])
        for row in duplicate_rows
    )


@pytest.mark.django_db
def test_import_preview_matches_existing_entity_by_composite_identity(
    client, users, composite_schema, composite_record
):
    workbook = workbook_from_rows(
        [
            ["公司编码", "员工号", "姓名", "状态", "valid_from"],
            ["C01", "E001", "张伟", "停缴", "2024-08-01"],
            ["C02", "E001", "李娜", "在缴", "2024-08-01"],
        ]
    )

    response = auth(client, users["editor"]).post(
        f"/api/v1/schemas/{composite_schema.id}/import/preview",
        {"file": workbook, "at": "2024-08-01", "missing_policy": "keep"},
        format="multipart",
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["summary"] == {
        "create": 1,
        "update": 1,
        "missing": 0,
        "invalid": 0,
        "unchanged": 0,
    }
    assert [row["business_code"] for row in payload["rows"]] == ["C01|E001", "C02|E001"]
    assert [row["display_code"] for row in payload["rows"]] == ["C01 / E001", "C02 / E001"]
    assert payload["rows"][0]["entity_id"] == composite_record.id
    assert payload["rows"][0]["action"] == "update"
    assert payload["rows"][0]["data_after"][IDENTITY_CODE_FIELD_KEY] == "C01|E001"


@pytest.mark.django_db
def test_changeset_create_entry_generates_composite_identity_code(
    client, users, composite_schema
):
    draft_response = auth(client, users["editor"]).post(
        f"/api/v1/schemas/{composite_schema.id}/changesets/",
        {"summary": "手工新增社保人员"},
        format="json",
    )
    draft = draft_response.json()

    response = client.post(
        f"/api/v1/changesets/{draft['id']}/entries/",
        {
            "action": "create",
            "valid_from": "2024-08-01",
            "data_after": {
                "company_code": "C01",
                "employee_no": "E001",
                "name": "张伟",
                "status": "在缴",
            },
        },
        format="json",
    )

    assert response.status_code == 201
    payload = response.json()
    entity = Entity.objects.get(schema=composite_schema, business_code="C01|E001")
    assert payload["entity_id"] == entity.id
    assert payload["business_code"] == "C01|E001"
    assert payload["data_after"][IDENTITY_CODE_FIELD_KEY] == "C01|E001"


@pytest.mark.django_db
def test_composite_identity_read_paths_return_display_code(
    client, users, composite_schema, composite_record
):
    api = auth(client, users["editor"])

    records_response = api.get(
        f"/api/v1/schemas/{composite_schema.id}/records/",
        {"at": "2024-06-02"},
    )
    timeline_response = api.get(f"/api/v1/entities/{composite_record.id}/timeline/")
    changeset_id = ChangeSet.objects.get(schema=composite_schema).id
    changeset_response = api.get(
        f"/api/v1/schemas/{composite_schema.id}/changesets/{changeset_id}/"
    )

    assert records_response.status_code == 200, records_response.json()
    record = records_response.json()["results"][0]
    assert record["business_code"] == "C01|E001"
    assert record["display_code"] == "C01 / E001"

    assert timeline_response.status_code == 200, timeline_response.json()
    entity = timeline_response.json()["entity"]
    assert entity["business_code"] == "C01|E001"
    assert entity["display_code"] == "C01 / E001"

    assert changeset_response.status_code == 200, changeset_response.json()
    entry = changeset_response.json()["entries"][0]
    assert entry["business_code"] == "C01|E001"
    assert entry["display_code"] == "C01 / E001"


@pytest.mark.django_db
def test_identity_display_template_is_used_by_read_import_and_export_paths(
    client, users, composite_schema, composite_record
):
    set_identity_display_template(composite_schema, "{employee_no} / {name}")
    api = auth(client, users["editor"])

    records_response = api.get(
        f"/api/v1/schemas/{composite_schema.id}/records/",
        {"at": "2024-06-02"},
    )
    timeline_response = api.get(f"/api/v1/entities/{composite_record.id}/timeline/")
    changeset_id = ChangeSet.objects.get(schema=composite_schema).id
    changeset_response = api.get(
        f"/api/v1/schemas/{composite_schema.id}/changesets/{changeset_id}/"
    )
    import_response = api.post(
        f"/api/v1/schemas/{composite_schema.id}/import/preview",
        {
            "file": workbook_from_rows(
                [
                    ["公司编码", "员工号", "姓名", "状态", "valid_from"],
                    ["C01", "E001", "张伟", "停缴", "2024-08-01"],
                    ["C02", "E001", "李娜", "在缴", "2024-08-01"],
                ]
            ),
            "at": "2024-08-01",
            "missing_policy": "keep",
        },
        format="multipart",
    )
    export_response = api.get(
        f"/api/v1/schemas/{composite_schema.id}/export/current",
        {"format": "csv", "at": "2024-06-02"},
    )

    assert records_response.status_code == 200, records_response.json()
    assert records_response.json()["results"][0]["display_code"] == "E001 / 张伟"
    assert timeline_response.status_code == 200, timeline_response.json()
    assert timeline_response.json()["entity"]["display_code"] == "E001 / 张伟"
    assert changeset_response.status_code == 200, changeset_response.json()
    assert changeset_response.json()["entries"][0]["display_code"] == "E001 / 张伟"
    assert import_response.status_code == 200, import_response.json()
    assert [row["display_code"] for row in import_response.json()["rows"]] == [
        "E001 / 张伟",
        "E001 / 李娜",
    ]
    assert export_response.status_code == 200, export_response.content.decode()
    csv_text = export_response.content.decode("utf-8-sig")
    assert "E001 / 张伟,2024-06-01" in csv_text
    assert "C01 / E001" not in csv_text


@pytest.mark.django_db
def test_identity_display_template_missing_field_does_not_fallback_to_business_code(
    client, users, composite_schema, composite_record
):
    fields_config = list(composite_schema.fields_config)
    fields_config.insert(-1, {"key": "nickname", "label": "昵称", "type": "text"})
    composite_schema.fields_config = fields_config
    composite_schema.save(update_fields=["fields_config"])
    set_identity_display_template(composite_schema, "{employee_no} / {nickname}")
    api = auth(client, users["editor"])

    records_response = api.get(
        f"/api/v1/schemas/{composite_schema.id}/records/",
        {"at": "2024-06-02"},
    )
    timeline_response = api.get(f"/api/v1/entities/{composite_record.id}/timeline/")
    changeset_id = ChangeSet.objects.get(schema=composite_schema).id
    changeset_response = api.get(
        f"/api/v1/schemas/{composite_schema.id}/changesets/{changeset_id}/"
    )
    export_response = api.get(
        f"/api/v1/schemas/{composite_schema.id}/export/current",
        {"format": "csv", "at": "2024-06-02"},
    )

    assert records_response.status_code == 200, records_response.json()
    assert records_response.json()["results"][0]["display_code"] == "E001 / —"
    assert timeline_response.status_code == 200, timeline_response.json()
    assert timeline_response.json()["entity"]["display_code"] == "E001 / —"
    assert changeset_response.status_code == 200, changeset_response.json()
    assert changeset_response.json()["entries"][0]["display_code"] == "E001 / —"
    assert export_response.status_code == 200, export_response.content.decode()
    csv_text = export_response.content.decode("utf-8-sig")
    assert "E001 / —,2024-06-01" in csv_text
    assert "C01|E001" not in csv_text


@pytest.mark.django_db
def test_composite_identity_exports_use_display_code(
    client, users, composite_schema, composite_record
):
    api = auth(client, users["editor"])

    current_response = api.get(
        f"/api/v1/schemas/{composite_schema.id}/export/current",
        {"format": "csv", "at": "2024-06-02"},
    )
    entity_response = api.get(f"/api/v1/entities/{composite_record.id}/export")

    assert current_response.status_code == 200, current_response.content.decode()
    csv_text = current_response.content.decode("utf-8-sig")
    assert csv_text.splitlines()[0].startswith("display_code,valid_from")
    assert "C01 / E001,2024-06-01" in csv_text
    assert "C01|E001" not in csv_text

    assert entity_response.status_code == 200
    workbook = load_workbook(BytesIO(entity_response.content), data_only=True)
    data_sheet = workbook["data"]
    metadata = {
        workbook["metadata"].cell(row=index, column=1).value: workbook["metadata"].cell(
            row=index, column=2
        ).value
        for index in range(1, workbook["metadata"].max_row + 1)
    }
    assert data_sheet.cell(row=1, column=1).value == "display_code"
    assert data_sheet.cell(row=2, column=1).value == "C01 / E001"
    assert metadata["display_code"] == "C01 / E001"
    assert metadata["business_code"] == "C01|E001"


@pytest.mark.django_db
def test_changeset_create_entry_rejects_empty_composite_identity_part(
    client, users, composite_schema
):
    draft_response = auth(client, users["editor"]).post(
        f"/api/v1/schemas/{composite_schema.id}/changesets/",
        {"summary": "手工新增社保人员"},
        format="json",
    )
    draft = draft_response.json()

    response = client.post(
        f"/api/v1/changesets/{draft['id']}/entries/",
        {
            "action": "create",
            "valid_from": "2024-08-01",
            "data_after": {"company_code": "C01", "employee_no": "", "name": "张伟"},
        },
        format="json",
    )

    assert response.status_code == 400
    assert "employee_no" in response.content.decode()


def composite_fields_config() -> list[dict]:
    return [
        {"key": "company_code", "label": "公司编码", "type": "text", "required": True},
        {"key": "employee_no", "label": "员工号", "type": "text", "required": True},
        {"key": "name", "label": "姓名", "type": "text"},
        {
            "key": "status",
            "label": "状态",
            "type": "enum",
            "validators": {"options": ["在缴", "停缴"]},
        },
        {
            "key": IDENTITY_CODE_FIELD_KEY,
            "label": "实体标识",
            "type": "text",
            "required": True,
            "indexed": True,
            "validators": {"max_length": 128},
            "hidden": True,
            "system": True,
            "identity_mode": "composite",
            "identity_field_keys": ["company_code", "employee_no"],
        },
    ]


def intake_fields_config(*, include_amount: bool = True) -> list[dict]:
    fields = [
        intake_field(1, "公司编码", "company_code", required=True),
        intake_field(2, "员工号", "employee_no", required=True),
        intake_field(3, "姓名", "name"),
    ]
    if include_amount:
        fields.append(
            {
                **intake_field(4, "社保基数", "social_base_amount"),
                "type": "number",
            }
        )
    return fields


def intake_field(source_index: int, label: str, key: str, *, required: bool = False) -> dict:
    return {
        "source_index": source_index,
        "source_column": label,
        "key": key,
        "label": label,
        "type": "text",
        "required": required,
        "indexed": required,
        "import": True,
        "validators": {},
    }


def hidden_identity_field(fields_config: list[dict]) -> dict:
    return next(field for field in fields_config if field["key"] == IDENTITY_CODE_FIELD_KEY)


def set_identity_display_template(schema: DataSchema, template: str) -> None:
    fields_config = list(schema.fields_config)
    hidden_field = hidden_identity_field(fields_config)
    hidden_field["identity_display_template"] = template
    schema.fields_config = fields_config
    schema.save(update_fields=["fields_config"])


def scan_token(client, user, workbook: BytesIO) -> str:
    response = auth(client, user).post(
        "/api/v1/excel-intake/scan",
        {"file": workbook},
        format="multipart",
    )
    assert response.status_code == 200
    return response.json()["upload_token"]


def workbook_from_rows(rows: list[list[object]]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    buffer.name = "social_base.xlsx"
    return buffer

from __future__ import annotations

import datetime as dt
import json
from collections import Counter
from typing import Any

from django.db import transaction
from openpyxl import load_workbook
from rest_framework.exceptions import ValidationError

from apps.changesets.api import get_changeset_payload
from apps.changesets.entry_editor import upsert_draft_entry
from apps.changesets.models import ChangeSet
from apps.schemas.identity import (
    IdentityResolutionError,
    field_is_system_hidden,
    resolve_business_code,
    resolve_display_code_or_fallback,
)
from apps.schemas.models import DataSchema
from apps.schemas.serializers import validation_issues
from apps.schemas.validation import FieldValidationError, validate_data_payload
from apps.temporal.models import TemporalRecord

from .coercion import coerce_value, parse_date
from .corrected import preview_from_corrected_payload
from .identity import build_identity_diagnostics, duplicate_identity_value_set

MAX_IMPORT_ROWS = 10_000


def preview_import(schema: DataSchema, file_obj, payload: dict) -> dict[str, Any]:
    at = parse_date(payload.get("at"), field="at")
    missing_policy = _missing_policy(payload.get("missing_policy"), schema)
    rows, mappings = _parse_workbook(schema, file_obj, at, payload)
    current = _current_rows(schema, at)
    identity_diagnostics = build_identity_diagnostics(
        schema.identity_field_key,
        schema.fields_config,
        rows,
    )
    duplicate_values = duplicate_identity_value_set(identity_diagnostics)
    seen_codes = {row["business_code"] for row in rows if row["business_code"]}
    preview_rows = [_preview_row(schema, row, current, duplicate_values) for row in rows]
    missing = _missing_rows(current, seen_codes, at, missing_policy)
    return {
        "schema_id": schema.id,
        "at": at.isoformat(),
        "missing_policy": missing_policy,
        "mappings": mappings,
        "summary": _summary(preview_rows, missing),
        "identity_diagnostics": identity_diagnostics,
        "rows": preview_rows,
        "missing": missing,
    }


def commit_import(schema: DataSchema, user: Any, file_obj, payload: dict) -> dict[str, Any]:
    preview = (
        preview_import(schema, file_obj, payload)
        if file_obj is not None
        else preview_from_corrected_payload(schema, payload)
    )
    invalid_rows = [row for row in preview["rows"] if row["action"] == "invalid"]
    if invalid_rows:
        raise ValidationError({"rows": "预览仍有校验失败行,请修正后再提交"})

    with transaction.atomic():
        change_set = ChangeSet.objects.create(
            schema=schema,
            summary=_summary_text(payload),
            status=ChangeSet.Status.DRAFT,
            approval_required=schema.approval_required,
            created_by=user,
            source=ChangeSet.Source.EXCEL,
        )
        for row in preview["rows"]:
            if row["action"] in {"create", "update"}:
                upsert_draft_entry(change_set, user, _entry_payload(row))
        for row in preview["missing"]:
            if row["action"] == "terminate":
                upsert_draft_entry(change_set, user, _entry_payload(row))
    return get_changeset_payload(schema, change_set.id)


def _parse_workbook(
    schema: DataSchema, file_obj, default_at: dt.date, payload: dict
) -> tuple[list[dict], list[dict]]:
    workbook = load_workbook(file_obj, data_only=True)
    sheet = workbook.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    mappings = _build_mappings(schema, headers, _mapping_overrides(payload))
    rows = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if _empty_row(row):
            continue
        if len(rows) >= MAX_IMPORT_ROWS:
            raise ValidationError({"file": f"单次导入不能超过 {MAX_IMPORT_ROWS} 行"})
        rows.append(_parse_row(schema, row_index, row, mappings, default_at))
    return rows, mappings


def _parse_row(
    schema: DataSchema, row_number: int, row: tuple, mappings: list[dict], default_at: dt.date
):
    data = {}
    valid_from = default_at
    field_lookup = {field["key"]: field for field in _active_fields(schema)}
    for column, mapping in enumerate(mappings):
        if column >= len(row) or not mapping["field_key"]:
            continue
        value = row[column]
        if mapping["field_key"] == "valid_from":
            valid_from = (
                parse_date(value, field="valid_from") if value not in (None, "") else default_at
            )
            continue
        field = field_lookup.get(mapping["field_key"])
        if field is None:
            raise ValidationError({"mappings_json": f"字段 {mapping['field_key']} 不存在"})
        coerced = coerce_value(field, value)
        if coerced not in (None, ""):
            data[field["key"]] = coerced
    business_code, identity_errors = _resolve_identity(schema, data)
    display_code = resolve_display_code_or_fallback(schema, data, business_code)
    return {
        "row_number": row_number,
        "business_code": business_code,
        "display_code": display_code,
        "valid_from": valid_from.isoformat(),
        "data_after": data,
        "errors": identity_errors,
    }


def _preview_row(
    schema: DataSchema,
    row: dict,
    current: dict[str, dict],
    duplicate_values: set[str],
) -> dict[str, Any]:
    errors = [*row.get("errors", []), *_row_errors(schema, row)]
    if row["business_code"] in duplicate_values:
        errors.append(_row_error(schema.identity_field_key, "duplicate_identity", "实体标识重复"))
    if errors:
        return {**row, "action": "invalid", "errors": errors, "changed_fields": []}
    current_row = current.get(row["business_code"])
    if current_row is None:
        return {
            **row,
            "action": "create",
            "errors": [],
            "changed_fields": sorted(row["data_after"]),
        }
    data_after = {**current_row["data_before"], **row["data_after"]}
    display_code = resolve_display_code_or_fallback(schema, data_after, row["business_code"])
    changed = _changed_fields(current_row["data_before"], data_after)
    action = "update" if changed else "unchanged"
    return {
        **row,
        "entity_id": current_row["entity_id"],
        "data_before": current_row["data_before"],
        "data_after": data_after,
        "display_code": display_code,
        "action": action,
        "errors": [],
        "changed_fields": changed,
    }


def _row_errors(schema: DataSchema, row: dict) -> list[dict]:
    has_required_error = any(error.get("code") == "required" for error in row.get("errors", []))
    if not row["business_code"]:
        return [] if has_required_error else [_row_error(schema.identity_field_key, "required", "实体标识必填")]
    try:
        validate_data_payload(schema.fields_config, row["data_after"])
    except FieldValidationError as exc:
        return validation_issues(exc)
    return []


def _row_error(field_key: str, code: str, message: str) -> dict:
    return {"path": f"data_payload.{field_key}", "code": code, "message": message}


def _current_rows(schema: DataSchema, at: dt.date) -> dict[str, dict]:
    records = (
        TemporalRecord.objects.select_related("entity")
        .filter(entity__schema=schema, is_superseded=False, valid_from__lte=at)
        .filter(valid_to__isnull=True)
        .order_by("entity__business_code", "-valid_from", "-id")
    )
    limited = {}
    for record in records:
        limited.setdefault(
            record.entity.business_code,
            {
                "entity_id": record.entity_id,
                "business_code": record.entity.business_code,
                "display_code": resolve_display_code_or_fallback(
                    schema, record.data_payload, record.entity.business_code
                ),
                "data_before": record.data_payload,
            },
        )
    ranged = (
        TemporalRecord.objects.select_related("entity")
        .filter(entity__schema=schema, is_superseded=False, valid_from__lte=at, valid_to__gt=at)
        .order_by("entity__business_code", "-valid_from", "-id")
    )
    for record in ranged:
        limited.setdefault(
            record.entity.business_code,
            {
                "entity_id": record.entity_id,
                "business_code": record.entity.business_code,
                "display_code": resolve_display_code_or_fallback(
                    schema, record.data_payload, record.entity.business_code
                ),
                "data_before": record.data_payload,
            },
        )
    return limited


def _missing_rows(
    current: dict[str, dict], seen_codes: set[str], at: dt.date, policy: str
) -> list[dict]:
    rows = []
    for business_code, item in sorted(current.items()):
        if business_code in seen_codes:
            continue
        rows.append(
            {
                "entity_id": item["entity_id"],
                "business_code": business_code,
                "display_code": item.get("display_code", business_code),
                "action": "terminate" if policy == "terminate" else "keep",
                "valid_from": at.isoformat(),
                "data_before": item["data_before"],
                "data_after": None,
                "errors": [],
                "changed_fields": [],
            }
        )
    return rows


def _build_mappings(
    schema: DataSchema, headers: list[str], overrides: dict[str, str]
) -> list[dict]:
    lookup = {}
    for field in _active_fields(schema):
        lookup[field["key"].lower()] = field["key"]
        lookup[field["label"].lower()] = field["key"]
    lookup["valid_from"] = "valid_from"
    valid_field_keys = {field["key"] for field in _active_fields(schema)} | {"valid_from", ""}
    invalid_field_keys = sorted(
        {field_key for field_key in overrides.values() if field_key not in valid_field_keys}
    )
    if invalid_field_keys:
        raise ValidationError({"mappings_json": f"字段 {invalid_field_keys[0]} 不存在"})
    return [
        {
            "source_column": header,
            "field_key": overrides.get(header, lookup.get(header.lower(), "")),
            "matched": header in overrides or header.lower() in lookup,
        }
        for header in headers
    ]


def _mapping_overrides(payload: dict) -> dict[str, str]:
    raw = payload.get("mappings_json") or payload.get("mappings")
    if not raw:
        return {}
    try:
        items = json.loads(raw) if isinstance(raw, str) else raw
    except json.JSONDecodeError as exc:
        raise ValidationError({"mappings_json": "必须是 JSON 数组"}) from exc
    if not isinstance(items, list):
        raise ValidationError({"mappings_json": "必须是 JSON 数组"})
    return {
        str(item.get("source_column", "")): str(item.get("field_key", ""))
        for item in items
        if isinstance(item, dict) and item.get("source_column")
    }


def _entry_payload(row: dict) -> dict[str, Any]:
    payload = {"action": row["action"], "valid_from": row["valid_from"]}
    if row["action"] == "create":
        payload["data_after"] = row["data_after"]
    elif row["action"] == "update":
        payload["entity_id"] = row["entity_id"]
        payload["data_after"] = row["data_after"]
    elif row["action"] == "terminate":
        payload["entity_id"] = row["entity_id"]
    return payload


def _summary(rows: list[dict], missing: list[dict]) -> dict[str, int]:
    counts = Counter(row["action"] for row in rows)
    missing_count = len(missing)
    return {
        "create": counts.get("create", 0),
        "update": counts.get("update", 0),
        "missing": missing_count,
        "invalid": counts.get("invalid", 0),
        "unchanged": counts.get("unchanged", 0),
    }


def _active_fields(schema: DataSchema) -> list[dict]:
    return [
        field
        for field in schema.fields_config
        if not field.get("deprecated", False) and not field_is_system_hidden(field)
    ]


def _resolve_identity(schema: DataSchema, data: dict[str, Any]) -> tuple[str, list[dict]]:
    try:
        return resolve_business_code(schema, data), []
    except IdentityResolutionError as exc:
        return "", [_row_error(exc.field_key, exc.code, exc.message)]


def _changed_fields(before: dict, after: dict) -> list[str]:
    keys = sorted(set(before) | set(after))
    return [key for key in keys if before.get(key) != after.get(key)]


def _missing_policy(value: object, schema: DataSchema) -> str:
    if value in (None, ""):
        return "keep" if schema.temporal_mode == DataSchema.TemporalMode.CONTINUOUS else "terminate"
    if value not in {"keep", "terminate"}:
        raise ValidationError({"missing_policy": "必须是 keep 或 terminate"})
    return str(value)


def _empty_row(row: tuple) -> bool:
    return all(value in (None, "") for value in row)


def _summary_text(payload: dict) -> str:
    value = payload.get("summary", "Excel 导入草稿")
    if not isinstance(value, str) or not value.strip():
        raise ValidationError({"summary": "必填"})
    return value.strip()[:200]

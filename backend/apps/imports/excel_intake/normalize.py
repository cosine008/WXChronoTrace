from __future__ import annotations

import copy
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from rest_framework.exceptions import ValidationError

from apps.imports.coercion import coerce_value, parse_date
from apps.imports.identity import build_identity_diagnostics, duplicate_identity_value_set
from apps.schemas.identity import (
    GENERATED_ENTITY_CODE_FIELD_KEY,
    IDENTITY_CODE_FIELD_KEY,
    IdentityResolutionError,
    auto_number_start_sequence,
    ensure_generated_entity_code_field,
    ensure_identity_code_field,
    format_auto_number_value,
    generated_entity_code_field,
    identity_auto_number_field,
    resolve_business_code,
    resolve_display_code_or_fallback,
    validate_composite_identity_keys,
)
from apps.schemas.serializers import validation_issues
from apps.schemas.validation import (
    FieldValidationError,
    validate_data_payload,
    validate_fields_config,
)

from .infer import infer_fields, key_from_label, mark_identity_candidates, schema_draft_from_payload

SOURCE_FIELDS = [
    {"key": "source_file", "label": "来源文件", "type": "text"},
    {"key": "source_sheet", "label": "来源 Sheet", "type": "text"},
    {"key": "source_row_no", "label": "来源行号", "type": "number"},
]
SYSTEM_CODE_FIELD = {
    "key": "system_code",
    "label": "系统编号",
    "type": "text",
    "required": True,
    "indexed": True,
    "validators": {},
}
PERSON_CODE_FIELD_KEY = "person_code"
PERSON_CODE_FIELD = {
    "key": PERSON_CODE_FIELD_KEY,
    "label": "人员编码",
    "type": "text",
    "required": True,
    "indexed": True,
    "validators": {"max_length": 32},
}


def build_preview(filename: str, content: bytes, payload: dict) -> dict[str, Any]:
    plan = import_plan_from_payload(payload)
    sheet_data = read_sheet_data(content, plan)
    user_fields = _user_fields(payload, sheet_data)
    schema_fields = schema_fields_from_drafts(user_fields, payload, plan)
    schema_draft = schema_draft_from_payload(payload, plan["sheet_name"], schema_fields)
    _validate_system_identity_policy(schema_draft)
    _ensure_identity(schema_draft, schema_fields)
    rows, identity_diagnostics = _preview_rows(sheet_data, user_fields, schema_draft, plan, filename)
    summary = _summary(rows)
    return {
        "schema_draft": schema_draft,
        "fields": mark_identity_candidates(user_fields),
        "import_plan": plan,
        "summary": summary,
        "identity_diagnostics": identity_diagnostics,
        "identity_warnings": _identity_warnings(schema_draft),
        "rows": rows,
        "errors": [error for row in rows for error in row["errors"]],
    }


def import_plan_from_payload(payload: dict) -> dict[str, Any]:
    sheet_name = payload.get("sheet_name")
    if not isinstance(sheet_name, str) or not sheet_name.strip():
        raise ValidationError({"sheet_name": "必填"})
    header_row = _positive_int(payload.get("header_row"), "header_row")
    data_start_row = _positive_int(payload.get("data_start_row"), "data_start_row")
    if data_start_row <= header_row:
        raise ValidationError({"data_start_row": "必须晚于 header_row"})
    missing_policy = payload.get("missing_policy", "keep")
    if missing_policy not in {"keep", "terminate"}:
        raise ValidationError({"missing_policy": "必须是 keep 或 terminate"})
    return {
        "sheet_name": sheet_name.strip(),
        "header_row": header_row,
        "data_start_row": data_start_row,
        "valid_from": parse_date(payload.get("valid_from"), field="valid_from").isoformat(),
        "missing_policy": missing_policy,
        "source_tracking": payload.get("source_tracking", True) is not False,
    }


def read_sheet_data(content: bytes, plan: dict) -> dict[str, Any]:
    workbook = _load_workbook(content)
    sheet_name = plan["sheet_name"]
    if sheet_name not in workbook.sheetnames:
        raise ValidationError({"sheet_name": "Sheet 不存在"})
    sheet = workbook[sheet_name]
    headers = _header_values(sheet, plan["header_row"])
    if not any(headers):
        raise ValidationError({"header_row": "表头行不能为空"})
    data_rows = _data_rows(sheet, plan["data_start_row"])
    if not data_rows and _has_data_after_header(sheet, plan["header_row"]):
        raise ValidationError({"data_start_row": "没有有效数据行"})
    return {"sheet_name": sheet_name, "headers": headers, "rows": data_rows}


def schema_fields_from_drafts(fields: list[dict], payload: dict, plan: dict) -> list[dict]:
    schema_fields = [_schema_field(field) for field in fields if field.get("import", True)]
    schema = payload.get("schema") if isinstance(payload.get("schema"), dict) else {}
    if schema.get("identity_mode") == "composite":
        schema_fields = _ensure_composite_identity_field(schema_fields, schema)
    if schema.get("identity_field_key") == GENERATED_ENTITY_CODE_FIELD_KEY:
        schema_fields = _ensure_generated_entity_code_field(schema_fields, schema, plan)
    if schema.get("identity_field_key") == "system_code":
        schema_fields = _ensure_system_code_field(schema_fields)
    if schema.get("identity_field_key") == PERSON_CODE_FIELD_KEY:
        schema_fields = _ensure_person_code_field(schema_fields)
    if plan["source_tracking"]:
        schema_fields = _ensure_source_fields(schema_fields)
    try:
        return validate_fields_config(schema_fields)
    except FieldValidationError as exc:
        raise ValidationError({"fields_config": validation_issues(exc)}) from exc


def _preview_rows(sheet_data: dict, fields: list[dict], schema_draft: dict, plan: dict, filename: str):
    pending_rows = []
    for offset, row in enumerate(sheet_data["rows"], start=1):
        data_after, errors, business_code, display_code = _row_payload(
            fields, row, schema_draft, plan, filename, offset
        )
        pending_rows.append(
            {
                "row_number": row["row_number"],
                "business_code": business_code,
                "display_code": display_code,
                "data_after": data_after,
                "errors": errors,
            }
        )

    identity_diagnostics = build_identity_diagnostics(
        schema_draft["identity_field_key"],
        schema_draft["fields_config"],
        pending_rows,
    )
    duplicate_values = duplicate_identity_value_set(identity_diagnostics)
    rows = []
    for item in pending_rows:
        errors = list(item["errors"])
        if item["business_code"] in duplicate_values:
            errors.append(
                _row_error(schema_draft["identity_field_key"], "duplicate_identity", "实体标识重复")
            )
        errors.extend(_validation_errors(schema_draft["fields_config"], item["data_after"]))
        rows.append(
            _preview_row(
                item["row_number"],
                item["business_code"],
                item["display_code"],
                plan["valid_from"],
                item["data_after"],
                errors,
            )
        )
    return rows, identity_diagnostics


def _row_payload(
    fields: list[dict], row: dict, schema_draft: dict, plan: dict, filename: str, offset: int
):
    data_after: dict[str, Any] = {}
    errors = []
    for field in fields:
        if not field.get("import", True):
            continue
        value = _source_value(row["values"], field.get("source_index"))
        coerced, error = _coerce_field(field, value)
        if error is not None:
            errors.append(error)
        if coerced not in (None, ""):
            data_after[field["key"]] = coerced
    if schema_draft["identity_field_key"] == "system_code":
        data_after["system_code"] = f"{schema_draft['schema_code']}-ROW-{offset:06d}"
    if schema_draft["identity_field_key"] == PERSON_CODE_FIELD_KEY and not data_after.get(
        PERSON_CODE_FIELD_KEY
    ):
        data_after[PERSON_CODE_FIELD_KEY] = _generated_person_code(plan["valid_from"], offset)
    auto_number_field = identity_auto_number_field(schema_draft)
    if auto_number_field and not data_after.get(schema_draft["identity_field_key"]):
        sequence = auto_number_start_sequence(auto_number_field) + offset - 1
        data_after[schema_draft["identity_field_key"]] = format_auto_number_value(
            schema_draft["schema_code"], auto_number_field, sequence, plan["valid_from"]
        )
    try:
        business_code = resolve_business_code(schema_draft, data_after)
    except IdentityResolutionError as exc:
        business_code = ""
        errors.append(_row_error(exc.field_key, exc.code, exc.message))
    display_code = resolve_display_code_or_fallback(schema_draft, data_after, business_code)
    if plan["source_tracking"]:
        data_after.update(
            {"source_file": filename, "source_sheet": plan["sheet_name"], "source_row_no": row["row_number"]}
        )
    return data_after, errors, business_code, display_code


def _user_fields(payload: dict, sheet_data: dict) -> list[dict]:
    raw = payload.get("fields_config") or payload.get("fields")
    if raw is None:
        return infer_fields(sheet_data["headers"], sheet_data["rows"])
    if not isinstance(raw, list):
        raise ValidationError({"fields_config": "必须是数组"})
    return mark_identity_candidates([_field_draft(field) for field in raw])


def _field_draft(field: dict) -> dict:
    if not isinstance(field, dict):
        raise ValidationError({"fields_config": "字段配置必须是对象"})
    return {
        **copy.deepcopy(field),
        "required": field.get("required", False),
        "indexed": field.get("indexed", False),
        "import": field.get("import", True),
        "validators": field.get("validators", {}),
    }


def _schema_field(field: dict) -> dict:
    return {
        "key": field.get("key"),
        "label": field.get("label"),
        "type": field.get("type"),
        "required": field.get("required", False),
        "indexed": field.get("indexed", False),
        "validators": field.get("validators", {}),
    }


def _preview_row(
    row_number: int,
    business_code: str,
    display_code: str,
    valid_from: str,
    data_after: dict,
    errors,
):
    return {
        "row_number": row_number,
        "business_code": business_code,
        "display_code": display_code,
        "valid_from": valid_from,
        "data_after": data_after,
        "action": "invalid" if errors else "create",
        "errors": errors,
        "changed_fields": [] if errors else sorted(data_after),
    }


def _validation_errors(fields_config: list[dict], data_after: dict) -> list[dict]:
    try:
        validate_data_payload(fields_config, data_after)
    except FieldValidationError as exc:
        return validation_issues(exc)
    return []


def _coerce_field(field: dict, value: object) -> tuple[object, dict | None]:
    try:
        return coerce_value(field, value), None
    except ValidationError as exc:
        return "", _row_error(field["key"], "coerce", str(exc.detail))


def _load_workbook(content: bytes):
    try:
        return load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise ValidationError({"file": "Excel 文件无法解析"}) from exc


def _header_values(sheet, header_row: int) -> list[str]:
    values = [cell.value for cell in next(sheet.iter_rows(min_row=header_row, max_row=header_row))]
    return [str(value).strip() if value is not None else "" for value in values]


def _data_rows(sheet, data_start_row: int) -> list[dict]:
    rows = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        if all(value in (None, "") for value in values):
            continue
        rows.append({"row_number": row_number, "values": tuple(values)})
    return rows


def _has_data_after_header(sheet, header_row: int) -> bool:
    return bool(_data_rows(sheet, header_row + 1))


def _source_value(values: tuple, source_index: object) -> object:
    if source_index in (None, ""):
        return ""
    index = int(source_index) - 1
    return values[index] if 0 <= index < len(values) else ""


def _ensure_identity(schema_draft: dict, fields: list[dict]) -> None:
    keys = {field["key"] for field in fields}
    if schema_draft.get("identity_mode") == "composite":
        try:
            validate_composite_identity_keys(fields, schema_draft.get("identity_field_keys", []))
        except IdentityResolutionError as exc:
            raise ValidationError({exc.field_key: exc.message}) from exc
        if IDENTITY_CODE_FIELD_KEY not in keys:
            raise ValidationError({"identity_field_key": "系统实体标识字段必须存在"})
        return
    if schema_draft["identity_field_key"] not in keys:
        raise ValidationError({"identity_field_key": "实体标识字段必须存在"})


def _ensure_composite_identity_field(fields: list[dict], schema: dict) -> list[dict]:
    identity_keys = schema.get("identity_field_keys")
    if not isinstance(identity_keys, list):
        identity_keys = []
    try:
        validate_composite_identity_keys(fields, [str(item) for item in identity_keys])
    except IdentityResolutionError as exc:
        raise ValidationError({exc.field_key: exc.message}) from exc
    return ensure_identity_code_field(fields, [str(item) for item in identity_keys])


def _ensure_generated_entity_code_field(fields: list[dict], schema: dict, plan: dict) -> list[dict]:
    schema_code = schema.get("schema_code") or key_from_label(plan["sheet_name"], 1)
    validators = schema.get("entity_code_config")
    if isinstance(validators, dict):
        fields = [generated_entity_code_field(str(schema_code), validators), *fields]
    return ensure_generated_entity_code_field(fields, str(schema_code))


def _ensure_system_code_field(fields: list[dict]) -> list[dict]:
    if any(field["key"] == "system_code" for field in fields):
        return fields
    return [SYSTEM_CODE_FIELD, *fields]


def _ensure_person_code_field(fields: list[dict]) -> list[dict]:
    if any(field["key"] == PERSON_CODE_FIELD_KEY for field in fields):
        return fields
    return [PERSON_CODE_FIELD, *fields]


def _ensure_source_fields(fields: list[dict]) -> list[dict]:
    existing = {field["key"] for field in fields}
    return [*fields, *[{**field, "required": True, "indexed": False, "validators": {}} for field in SOURCE_FIELDS if field["key"] not in existing]]


def _positive_int(value: object, field: str) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError) as exc:
        raise ValidationError({field: "必须是正整数"}) from exc
    if parsed < 1:
        raise ValidationError({field: "必须是正整数"})
    return parsed


def _row_error(field_key: str, code: str, message: str) -> dict:
    return {"path": f"data_payload.{field_key}", "code": code, "message": message}


def _validate_system_identity_policy(schema_draft: dict) -> None:
    if (
        schema_draft.get("identity_field_key") == "system_code"
        and schema_draft.get("temporal_mode") == "periodic"
    ):
        raise ValidationError(
            {
                "identity_field_key": (
                    "行号型 system_code 只适合一次性新表导入；周期型表请使用员工号、"
                    "person_code 或组合标识。"
                )
            }
        )


def _identity_warnings(schema_draft: dict) -> list[dict[str, str]]:
    identity_key = schema_draft.get("identity_field_key")
    if identity_key == PERSON_CODE_FIELD_KEY:
        return [
            {
                "code": "person_code_generated",
                "message": "将为本次新表接入生成稳定 person_code，后续周期导入需继续携带该字段。",
            }
        ]
    if identity_auto_number_field(schema_draft):
        return [
            {
                "code": "entity_code_generated",
                "message": "将为本次新表接入生成唯一实体编码，后续导入需继续携带该字段。",
            }
        ]
    if identity_key == "system_code":
        return [
            {
                "code": "row_number_system_code",
                "message": "system_code 按导入行号生成，只适合一次性新表接入，不适合作为周期导入身份。",
            }
        ]
    return []


def _generated_person_code(valid_from: str, offset: int) -> str:
    return f"P-{valid_from[:4]}-{offset:06d}"


def _summary(rows: list[dict]) -> dict[str, int]:
    invalid = sum(1 for row in rows if row["action"] == "invalid")
    return {
        "create": len(rows) - invalid,
        "update": 0,
        "missing": 0,
        "invalid": invalid,
        "unchanged": 0,
    }

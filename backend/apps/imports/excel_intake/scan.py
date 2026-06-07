from __future__ import annotations

import datetime as dt
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from rest_framework.exceptions import ValidationError

from .storage import TOKEN_TTL_SECONDS, save_upload

PREVIEW_ROW_LIMIT = 25


def scan_upload(file_obj) -> dict[str, Any]:
    filename = getattr(file_obj, "name", "")
    content = file_obj.read()
    upload_token = save_upload(filename, content)
    workbook = _load_workbook(content)
    sheets = [_sheet_summary(sheet) for sheet in workbook.worksheets]
    if not sheets:
        raise ValidationError({"file": "未找到可读 Sheet"})
    return {
        "upload_token": upload_token,
        "expires_in_seconds": TOKEN_TTL_SECONDS,
        "filename": filename,
        "sheets": sheets,
    }


def _load_workbook(content: bytes):
    try:
        return load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise ValidationError({"file": "Excel 文件无法解析"}) from exc


def _sheet_summary(sheet: Worksheet) -> dict[str, Any]:
    row_count, column_count = _sheet_size(sheet)
    preview_rows = _preview_rows(sheet)
    header_row = _recommended_header_row(preview_rows)
    return {
        "name": sheet.title,
        "row_count": row_count,
        "column_count": column_count,
        "recommended_header_row": header_row,
        "recommended_data_start_row": header_row + 1,
        "preview_rows": preview_rows,
    }


def _sheet_size(sheet: Worksheet) -> tuple[int, int]:
    if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(1, 1).value is None:
        return 0, 0
    return sheet.max_row, sheet.max_column


def _preview_rows(sheet: Worksheet) -> list[list[Any]]:
    rows = []
    for row in sheet.iter_rows(max_row=PREVIEW_ROW_LIMIT, values_only=True):
        rows.append(_trim_empty_tail([_cell_value(value) for value in row]))
    return rows


def _recommended_header_row(rows: list[list[Any]]) -> int:
    for index, row in enumerate(rows, start=1):
        if _non_empty_count(row) >= 2:
            return index
    for index, row in enumerate(rows, start=1):
        if _non_empty_count(row) > 0:
            return index
    return 1


def _cell_value(value: object) -> object:
    if isinstance(value, dt.datetime):
        return value.isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return value


def _trim_empty_tail(row: list[Any]) -> list[Any]:
    while row and row[-1] in (None, ""):
        row.pop()
    return row


def _non_empty_count(row: list[Any]) -> int:
    return sum(1 for value in row if value not in (None, ""))

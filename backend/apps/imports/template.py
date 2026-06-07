from __future__ import annotations

import datetime as dt
from io import BytesIO

from openpyxl import Workbook
from openpyxl.comments import Comment

from apps.schemas.identity import field_is_system_hidden
from apps.schemas.models import DataSchema


def build_template_workbook(schema: DataSchema) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = schema.schema_code[:31]
    fields = _active_fields(schema)
    headers = [field["label"] for field in fields] + ["valid_from"]
    sheet.append(headers)
    for index, field in enumerate(fields, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.comment = Comment(_field_comment(field, schema.identity_field_key), "ChronoTrace")
        sheet.cell(row=2, column=index).value = _example_value(field)
        sheet.column_dimensions[cell.column_letter].width = max(14, len(str(cell.value)) + 4)
    valid_from = sheet.cell(row=1, column=len(headers))
    valid_from.comment = Comment(
        "Effective date, format YYYY-MM-DD; blank uses the import wizard date.",
        "ChronoTrace",
    )
    sheet.cell(row=2, column=len(headers)).value = dt.date.today().isoformat()
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _active_fields(schema: DataSchema) -> list[dict]:
    return [
        field
        for field in schema.fields_config
        if not field.get("deprecated", False) and not field_is_system_hidden(field)
    ]


def _field_comment(field: dict, identity_key: str) -> str:
    validators = field.get("validators") or {}
    lines = [f"field: {field['key']}", f"type: {field['type']}"]
    if field.get("required") or field["key"] == identity_key:
        lines.append("required")
    if field["type"] in {"attachment", "image"}:
        lines.append("Use uploaded asset id list, for example: 1,2")
    if field["type"] == "formula":
        lines.append("Computed field; leave blank in import files")
    if validators:
        lines.append(f"validators: {validators}")
    return "\n".join(lines)


def _example_value(field: dict) -> object:
    options = (field.get("validators") or {}).get("options")
    if options:
        return options[0]
    return {
        "text": "example text",
        "longtext": "example note",
        "markdown": "## Example\n\n- item",
        "number": 100,
        "date": dt.date.today().isoformat(),
        "datetime": dt.datetime.now(dt.UTC).isoformat(),
        "boolean": True,
        "multi-enum": "option1,option2",
        "person": 1,
        "reference": 1,
        "auto-number": "AUTO-001",
        "attachment": "1,2",
        "image": "1,2",
        "formula": "",
    }.get(field["type"], "")
